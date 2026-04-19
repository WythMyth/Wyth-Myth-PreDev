"""
views.py  (meetings app)
─────────────────────────
Only the views that changed are shown below. Copy-paste them over the
matching classes / functions in your existing views.py.
"""

import json
import uuid
from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.db.models import Sum
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from openpyxl import load_workbook

from poll.permission import PermissionRequiredMixin

from .forms import MeetingForm, RecordingForm, TagForm
from .models import ClassRecording, MeetingSchedule, Tag
from .notifications import schedule_meeting_notifications
from .utils import RecurrenceSpec, generate_occurrence_dates

User = get_user_model()


# ── Shared context helper ──────────────────────────────────────────────────


def _balance_context(user: object) -> dict:
    if user.is_superuser:
        total = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
    else:
        total = user.balance
    return {"total_balance": total}


def _parse_time(time_value):
    """Convert time value to datetime.time object."""
    from datetime import time as time_cls

    if time_value is None:
        return None
    if isinstance(time_value, time_cls):
        return time_value
    if isinstance(time_value, str):
        try:
            parts = time_value.split(":")
            if len(parts) >= 2:
                return time_cls(int(parts[0]), int(parts[1]))
        except (ValueError, IndexError):
            pass
    return None


# ═══════════════════════════════════════════════════════════════════════════
# MEETING VIEWS
# ═══════════════════════════════════════════════════════════════════════════


class MeetingListView(LoginRequiredMixin, ListView):
    model = MeetingSchedule
    template_name = "meetings/meeting_list.html"
    context_object_name = "meetings"
    paginate_by = 20

    def get_queryset(self):
        qs = MeetingSchedule.objects.all().order_by(
            "is_expired", "-date", "-start_time"
        )

        user = self.request.user
        # Only restrict non-superusers
        if not user.is_superuser:
            qs = qs.filter(guests__in=user.tags.all()).distinct()

        title = self.request.GET.get("title", "").strip()
        from_date = self.request.GET.get("from_date")
        to_date = self.request.GET.get("to_date")
        status = self.request.GET.get("status")

        if title:
            qs = qs.filter(title__icontains=title)
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
        if status == "active":
            qs = qs.filter(is_expired=False)
        elif status == "expired":
            qs = qs.filter(is_expired=True)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_balance_context(self.request.user))
        ctx["status_filter"] = self.request.GET.get("status", "")
        ctx["title_filter"] = self.request.GET.get("title", "")
        ctx["from_date"] = self.request.GET.get("from_date", "")
        ctx["to_date"] = self.request.GET.get("to_date", "")
        return ctx


class MeetingDetailView(LoginRequiredMixin, DetailView):
    model = MeetingSchedule
    template_name = "meetings/meeting_detail.html"
    context_object_name = "meeting"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["recordings"] = self.object.recording.all()
        ctx.update(_balance_context(self.request.user))
        return ctx


class MeetingCreateView(LoginRequiredMixin,PermissionRequiredMixin, View):
    """
    Handles both single and recurring meetings.
    Recurring: generates multiple MeetingSchedule objects sharing a series_id.
    """

    template_name = "meetings/meeting_form.html"
    permission_flags = ["is_superuser"]

    def get(self, request, *args, **kwargs):
        form = MeetingForm()
        ctx = {"form": form}
        ctx.update(_balance_context(request.user))
        return render(request, self.template_name, ctx)

    def post(self, request, *args, **kwargs):
        form = MeetingForm(request.POST)
        if not form.is_valid():
            ctx = {"form": form}
            ctx.update(_balance_context(request.user))
            return render(request, self.template_name, ctx)

        cd = form.cleaned_data

        with transaction.atomic():
            series_id = uuid.uuid4()

            # ── Determine dates to create ──────────────────────────────────
            if cd.get("is_recurring"):
                spec = RecurrenceSpec(
                    start_date=cd["date"],
                    recurrence_type=cd["recurrence_type"],
                    interval=cd["interval"],
                    days_of_week=[int(x) for x in cd.get("days_of_week", [])] or None,
                    monthly_mode=cd.get("monthly_mode") or None,
                    monthly_day_of_month=cd.get("monthly_day_of_month") or None,
                    monthly_nth=(
                        int(cd["monthly_nth"]) if cd.get("monthly_nth") else None
                    ),
                    monthly_weekday=(
                        int(cd["monthly_weekday"])
                        if cd.get("monthly_weekday") not in (None, "")
                        else None
                    ),
                    yearly_mode=cd.get("yearly_mode") or None,
                    yearly_month=(
                        int(cd["yearly_month"]) if cd.get("yearly_month") else None
                    ),
                    yearly_day_of_month=cd.get("yearly_day_of_month") or None,
                    yearly_nth=int(cd["yearly_nth"]) if cd.get("yearly_nth") else None,
                    yearly_weekday=(
                        int(cd["yearly_weekday"])
                        if cd.get("yearly_weekday") not in (None, "")
                        else None
                    ),
                    end_mode=cd.get("end_mode"),
                    end_on_date=cd.get("end_on_date"),
                    end_after_occurrences=cd.get("end_after_occurrences"),
                )
                dates = generate_occurrence_dates(spec)
                if not dates:
                    messages.error(
                        request, "No meetings could be generated with those settings."
                    )
                    ctx = {"form": form}
                    ctx.update(_balance_context(request.user))
                    return render(request, self.template_name, ctx)
            else:
                dates = [cd["date"]]

            # ── Common field values ────────────────────────────────────────
            common = dict(
                meeting_user=request.user,
                title=cd["title"],
                description=cd.get("description", ""),
                meeting_url=cd.get("meeting_url", ""),
                password=cd.get("password", ""),
                start_time=cd["start_time"],
                end_time=cd.get("end_time"),
                is_sms=cd.get("is_sms", False),
                enable_all_email_notification=cd.get(
                    "enable_all_email_notification", False
                ),
                notice_3_weeks=cd.get("notice_3_weeks", False),
                notice_2_weeks=cd.get("notice_2_weeks", False),
                notice_1_week=cd.get("notice_1_week", False),
                notice_1_day=cd.get("notice_1_day", False),
                notice_10_min=cd.get("notice_10_min", False),
                series_id=series_id,
                is_recurring=False,  # individual occurrences are not recurring
            )

            # ── Create one MeetingSchedule per date ────────────────────────
            created_meetings = []
            for idx, d in enumerate(dates, start=1):
                meeting = MeetingSchedule.objects.create(
                    date=d,
                    occurrence_index=idx,
                    **common,
                )
                created_meetings.append(meeting)

            # Mark the first occurrence as the series root
            if cd.get("is_recurring") and created_meetings:
                first = created_meetings[0]
                first.is_recurring = True
                first.save(update_fields=["is_recurring"])

            # ── Handle tags for all meetings ────────────────────────────────
            tags = cd.get("guests")
            if tags:
                for m in created_meetings:
                    m.guests.set(tags)

            # ── Schedule email notifications for every occurrence ──────────
            for m in created_meetings:
                schedule_meeting_notifications(m)

        messages.success(
            request, f"{len(created_meetings)} meeting(s) created successfully."
        )
        return redirect("meeting-list")


class MeetingUpdateView(LoginRequiredMixin, UpdateView):
    """
    Edits a single MeetingSchedule occurrence (not the whole series).
    Re-schedules notifications after save.
    """

    model = MeetingSchedule
    form_class = MeetingForm
    template_name = "meetings/meeting_form.html"
    success_url = reverse_lazy("meeting-list")

    def form_valid(self, form):
        # Save first
        response = super().form_valid(form)
        # Re-schedule notifications with updated values
        schedule_meeting_notifications(self.object)
        messages.success(self.request, "Meeting updated successfully!")
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_balance_context(self.request.user))
        return ctx


class MeetingDeleteView(LoginRequiredMixin, DeleteView):
    model = MeetingSchedule
    template_name = "meetings/meeting_confirm_delete.html"
    success_url = reverse_lazy("meeting-list")

    def delete(self, request, *args, **kwargs):
        meeting = self.get_object()
        meeting_title = meeting.title
        meeting_date = meeting.date
        meeting_time = meeting.start_time
        deleted_by = request.user.get_full_name() or request.user.username

        if request.user.is_superuser:
            users = User.objects.filter(email__isnull=False).exclude(email="")
            self._send_deletion_email(
                users, meeting_title, meeting_date, meeting_time, deleted_by
            )

        messages.success(request, "Meeting deleted successfully!")
        return super().delete(request, *args, **kwargs)

    def _send_deletion_email(self, users, title, date, time, deleted_by):
        subject = f"Meeting Cancelled: {title}"
        ctx = {
            "meeting_title": title,
            "meeting_date": date,
            "meeting_time": time,
            "deleted_by": deleted_by,
        }
        try:
            html_content = render_to_string("emails/meeting_deleted_email.html", ctx)
        except Exception:
            html_content = None

        text_content = (
            f"Dear User,\n\nThe following meeting has been cancelled:\n\n"
            f"Title:        {title}\nDate:         {date}\n"
            f"Time:         {time}\nCancelled by: {deleted_by}\n\n"
            "We apologise for any inconvenience.\n\nMeeting Management Team"
        )
        to_emails = [u.email for u in users if u.email]
        if not to_emails:
            return
        try:
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_content,
                from_email=None,
                to=to_emails,
            )
            if html_content:
                email.attach_alternative(html_content, "text/html")
            email.send(fail_silently=False)
        except Exception as e:
            print(f"Error sending deletion notification: {e}")


class MeetingCopyView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        original = get_object_or_404(MeetingSchedule, pk=pk)
        copied = MeetingSchedule.objects.create(
            meeting_user=request.user,
            title=f"{original.title} (Copy)",
            description=original.description,
            date=original.date,
            start_time=original.start_time,
            end_time=original.end_time,
            meeting_url=original.meeting_url,
            password=original.password,
            is_sms=original.is_sms,
            series_id=uuid.uuid4(),
            occurrence_index=1,
            # Copy notification settings
            enable_all_email_notification=original.enable_all_email_notification,
            notice_3_weeks=original.notice_3_weeks,
            notice_2_weeks=original.notice_2_weeks,
            notice_1_week=original.notice_1_week,
            notice_1_day=original.notice_1_day,
            notice_10_min=original.notice_10_min,
        )
        schedule_meeting_notifications(copied)
        return JsonResponse({"status": "success", "copied_id": copied.pk})


# ═══════════════════════════════════════════════════════════════════════════
# RECORDING VIEWS  (unchanged – copied for completeness)
# ═══════════════════════════════════════════════════════════════════════════


class RecordingListView(LoginRequiredMixin, ListView):
    model = ClassRecording
    template_name = "recordings/recording_list.html"
    context_object_name = "recordings"
    paginate_by = 20

    def get_queryset(self):
        qs = ClassRecording.objects.select_related("meeting").order_by("-meeting__date")
        user = self.request.user
        # Apply tag restriction
        if not user.is_superuser:
            qs = qs.filter(meeting__guests__in=user.tags.all()).distinct()
        title = self.request.GET.get("title", "").strip()
        from_date = self.request.GET.get("from_date", "").strip()
        to_date = self.request.GET.get("to_date", "").strip()
        if title:
            qs = qs.filter(meeting__title__icontains=title)
        if from_date:
            qs = qs.filter(meeting__date__gte=from_date)
        if to_date:
            qs = qs.filter(meeting__date__lte=to_date)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_balance_context(self.request.user))
        ctx["title_filter"] = self.request.GET.get("title", "")
        ctx["from_date_filter"] = self.request.GET.get("from_date", "")
        ctx["to_date_filter"] = self.request.GET.get("to_date", "")
        return ctx


class RecordingDetailView(LoginRequiredMixin, DetailView):
    model = ClassRecording
    template_name = "recordings/recording_detail.html"
    context_object_name = "recording"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_balance_context(self.request.user))
        return ctx


class RecordingCreateView(LoginRequiredMixin, CreateView):
    model = ClassRecording
    form_class = RecordingForm
    template_name = "recordings/recording_form.html"
    success_url = reverse_lazy("recording-list")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["meeting"].queryset = MeetingSchedule.objects.filter(
            meeting_user=self.request.user, is_expired=False
        )
        return form

    def form_valid(self, form):
        messages.success(self.request, "Recording created successfully!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_balance_context(self.request.user))
        return ctx


class RecordingUpdateView(LoginRequiredMixin, UpdateView):
    model = ClassRecording
    form_class = RecordingForm
    template_name = "recordings/recording_form.html"
    success_url = reverse_lazy("recording-list")

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["meeting"].queryset = MeetingSchedule.objects.filter(
            meeting_user=self.request.user
        )
        return form

    def form_valid(self, form):
        messages.success(self.request, "Recording updated successfully!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(_balance_context(self.request.user))
        return ctx


class RecordingDeleteView(LoginRequiredMixin, DeleteView):
    model = ClassRecording
    template_name = "recordings/recording_confirm_delete.html"
    success_url = reverse_lazy("recording-list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Recording deleted successfully!")
        return super().delete(request, *args, **kwargs)


# ═══════════════════════════════════════════════════════════════════════════
# UPLOAD / DOWNLOAD / CALENDAR  (unchanged)
# ═══════════════════════════════════════════════════════════════════════════


@login_required
def meeting_upload(request):
    template = "meetings/meeting_upload.html"

    if request.method == "POST" and "file" in request.FILES:
        file = request.FILES["file"]
        if file.name.endswith(".xlsx"):
            try:
                wb = load_workbook(file)
                sheet = wb.active
                success_count = error_count = 0
                errors = []

                for row in range(2, sheet.max_row + 1):
                    try:
                        title = sheet.cell(row=row, column=1).value
                        description = sheet.cell(row=row, column=2).value
                        date_value = sheet.cell(row=row, column=3).value
                        start_time = str(sheet.cell(row=row, column=4).value)
                        end_time = str(sheet.cell(row=row, column=5).value)
                        meeting_url = sheet.cell(row=row, column=6).value
                        password = sheet.cell(row=row, column=7).value

                        is_sms = False
                        if sheet.max_column >= 8:
                            v = sheet.cell(row=row, column=8).value
                            if isinstance(v, bool):
                                is_sms = v
                            elif isinstance(v, str):
                                is_sms = v.lower() in ["true", "yes", "1"]

                        if not title:
                            continue

                        MeetingSchedule.objects.filter(
                            meeting_user=request.user, title=title, date=date_value
                        ).delete()

                        MeetingSchedule.objects.create(
                            meeting_user=request.user,
                            title=title,
                            description=description,
                            date=date_value,
                            start_time=start_time,
                            end_time=end_time,
                            meeting_url=meeting_url,
                            password=password,
                            is_sms=is_sms,
                        )
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row}: {e}")

                if success_count:
                    messages.success(request, f"{success_count} meetings uploaded.")
                if error_count:
                    messages.error(
                        request, f"{error_count} failed. {errors[0] if errors else ''}"
                    )
                return redirect("meeting-list")
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
        else:
            messages.error(request, "Please upload a valid .xlsx file!")

    ctx = _balance_context(request.user)
    return render(request, template, ctx)


@login_required
def download_meeting_template(request):
    from django.http import HttpResponse
    from openpyxl import Workbook

    wb, ws = Workbook(), None
    wb.active.title = "Meetings"
    ws = wb.active
    headers = [
        "Title",
        "Description",
        "Date (YYYY-MM-DD)",
        "Start Time",
        "End Time",
        "Meeting URL",
        "Password",
        "Send SMS (True/False)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    example = [
        "Weekly Team Meeting",
        "Regular sync-up",
        "2025-05-27",
        "10:00",
        "11:00",
        "https://example.com/meeting",
        "pass123",
        "False",
    ]
    for col, v in enumerate(example, 1):
        ws.cell(row=2, column=col).value = v

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="meeting_upload_template.xlsx"'
    )
    wb.save(response)
    return response


class CalendarView(LoginRequiredMixin, View):
    template_name = "meetings/calendar.html"

    def get(self, request, *args, **kwargs):
        today = timezone.localdate()
        meetings = MeetingSchedule.objects.filter(date__gte=today).order_by(
            "date", "start_time"
        )
        user = request.user
         #  Apply tag-based restriction
        if not user.is_superuser:
            meetings = meetings.filter(
                guests__in=user.tags.all()
            ).distinct()
        event_data = []
        for meeting in meetings:
            try:
                # Parse times safely
                start_time = _parse_time(meeting.start_time)
                end_time = _parse_time(meeting.end_time)

                if not start_time or not end_time or not meeting.date:
                    continue

                start_dt = datetime.combine(meeting.date, start_time)
                end_dt = datetime.combine(meeting.date, end_time)
                if timezone.is_naive(start_dt):
                    start_dt = timezone.make_aware(start_dt)
                if timezone.is_naive(end_dt):
                    end_dt = timezone.make_aware(end_dt)
                if end_dt < start_dt:
                    end_dt += timedelta(days=1)
                event_data.append(
                    {
                        "id": meeting.id,
                        "title": meeting.title,
                        "start": start_dt.isoformat(),
                        "end": end_dt.isoformat(),
                    }
                )
            except Exception as e:
                print(f"Error processing meeting {meeting.id}: {e}")

        ctx = {"event_data": json.dumps(event_data)}
        ctx.update(_balance_context(request.user))
        return render(request, self.template_name, ctx)


# -------------------- Tag --------------------


class TagListView(LoginRequiredMixin, ListView):
    model = Tag
    template_name = "dashboard/settings/tag_list.html"
    context_object_name = "tags"
    paginate_by = 50
    ordering = [
        "order",
    ]


class TagCreateView(LoginRequiredMixin, CreateView):
    model = Tag
    form_class = TagForm
    template_name = "dashboard/settings/tag_form.html"
    success_url = reverse_lazy("tag_list")

    def form_invalid(self, form):
        print(form.errors)
        return super().form_invalid(form)


class TagUpdateView(LoginRequiredMixin, UpdateView):
    model = Tag
    form_class = TagForm
    template_name = "dashboard/settings/tag_form.html"
    success_url = reverse_lazy("tag_list")


class TagDeleteView(LoginRequiredMixin, View):
    def delete(self, request, *args, **kwargs):
        tag = get_object_or_404(Tag, pk=kwargs["pk"])
        tag.delete()
        return JsonResponse({"success": True})


# from datetime import datetime, timedelta
# import json
# from django.http import JsonResponse
# from datetime import timedelta
# from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
# from django.contrib import messages
# from django.contrib.auth.decorators import login_required
# from django.contrib.auth.mixins import LoginRequiredMixin
# from django.db.models import Sum
# from django.shortcuts import get_object_or_404, redirect, render
# from django.urls import reverse_lazy
# from django.utils import timezone
# from django.utils.timezone import now
# from django.views import View
# from django.views.generic import (
#     CreateView,
#     DeleteView,
#     DetailView,
#     ListView,
#     TemplateView,
#     UpdateView,
# )
# from django.utils.timezone import localtime, is_naive, make_aware
# from openpyxl import load_workbook
# from django.core.mail import EmailMultiAlternatives
# from django.conf import settings
# from django.template.loader import render_to_string
# from accounts.models import User

# from .forms import MeetingForm, RecordingForm
# from .models import ClassRecording, MeetingSchedule

# class MeetingListView(LoginRequiredMixin, ListView):
#     model = MeetingSchedule
#     template_name = "meetings/meeting_list.html"
#     context_object_name = "meetings"
#     paginate_by = 20
#     def get_queryset(self):
#         today = timezone.localtime().date()
#         qs = MeetingSchedule.objects.all()

#         # Order latest date first
#         qs = qs.order_by('is_expired', '-date', '-start_time')

#         # Filters
#         title = self.request.GET.get('title', '').strip()
#         from_date = self.request.GET.get('from_date')
#         to_date = self.request.GET.get('to_date')
#         status = self.request.GET.get('status')

#         if title:
#             qs = qs.filter(title__icontains=title)

#         if from_date:
#             qs = qs.filter(date__gte=from_date)

#         if to_date:
#             qs = qs.filter(date__lte=to_date)

#         if status == "active":
#             qs = qs.filter(is_expired=False)
#         elif status == "expired":
#             qs = qs.filter(is_expired=True)

#         return qs

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)

#         # Filters in context for template
#         context["status_filter"] = self.request.GET.get("status", "")
#         context["title_filter"] = self.request.GET.get("title", "")
#         context["from_date"] = self.request.GET.get("from_date", "")
#         context["to_date"] = self.request.GET.get("to_date", "")

#         # ✅ Add balance if user is superuser
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance

#         return context

# class MeetingDetailView(LoginRequiredMixin, DetailView):
#     model = MeetingSchedule
#     template_name = "meetings/meeting_detail.html"
#     context_object_name = "meeting"

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context["recordings"] = self.object.recording.all()
#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
#         return context


# class MeetingCreateView(LoginRequiredMixin, CreateView):
#     model = MeetingSchedule
#     form_class = MeetingForm
#     template_name = "meetings/meeting_form.html"
#     success_url = reverse_lazy("meeting-list")

#     def form_valid(self, form):
#         form.instance.meeting_user = self.request.user
#         messages.success(self.request, "Meeting created successfully!")
#         return super().form_valid(form)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
#         return context


# class MeetingUpdateView(LoginRequiredMixin, UpdateView):
#     model = MeetingSchedule
#     form_class = MeetingForm
#     template_name = "meetings/meeting_form.html"
#     success_url = reverse_lazy("meeting-list")

#     def form_valid(self, form):
#         messages.success(self.request, "Meeting updated successfully!")
#         return super().form_valid(form)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
#         return context

# class MeetingDeleteView(LoginRequiredMixin, DeleteView):
#     model = MeetingSchedule
#     template_name = "meetings/meeting_confirm_delete.html"
#     success_url = reverse_lazy("meeting-list")

#     def delete(self, request, *args, **kwargs):
#         # Store meeting details before deletion
#         meeting = self.get_object()
#         meeting_title = meeting.title
#         meeting_date = meeting.date
#         meeting_time = meeting.start_time
#         deleted_by = request.user.get_full_name() or request.user.username

#         # Check if the user is a superuser
#         if request.user.is_superuser:
#             # Get all users with valid email addresses
#             users = User.objects.filter(email__isnull=False).exclude(email='')

#             # Send email to all users
#             self.send_deletion_notification_email(
#                 users, meeting_title, meeting_date, meeting_time, deleted_by
#             )

#         messages.success(request, "Meeting deleted successfully!")
#         return super().delete(request, *args, **kwargs)

#     def send_deletion_notification_email(self, users, meeting_title, meeting_date, meeting_time, deleted_by):
#         """Send HTML email notification to all users about meeting deletion"""
#         subject = f"Meeting Cancelled: {meeting_title}"

#         # Context data for the template
#         context = {
#             'meeting_title': meeting_title,
#             'meeting_date': meeting_date,
#             'meeting_time': meeting_time,
#             'deleted_by': deleted_by,
#         }

#         # Render HTML template
#         html_content = render_to_string('emails/meeting_deleted_email.html', context)

#         # Optional: Create a plain text version as fallback
#         text_content = f"""
# Dear User,

# We would like to inform you that the following meeting has been cancelled:

# Meeting Details:
# - Title: {meeting_title}
# - Date: {meeting_date}
# - Time: {meeting_time}
# - Cancelled by: {deleted_by}

# We apologize for any inconvenience this may cause.

# Best regards,
# Meeting Management Team
#         """

#         # Get recipient email addresses
#         recipient_emails = [user.email for user in users if user.email]

#         if recipient_emails:
#             try:
#                 # Create EmailMultiAlternatives instance
#                 email = EmailMultiAlternatives(
#                     subject=subject,
#                     body=text_content,  # Plain text version
#                     from_email=settings.DEFAULT_FROM_EMAIL,
#                     to=recipient_emails,
#                 )

#                 # Attach HTML version
#                 email.attach_alternative(html_content, "text/html")

#                 # Send the email
#                 email.send(fail_silently=False)

#                 print(f"Meeting deletion notification sent to {len(recipient_emails)} users")
#             except Exception as e:
#                 print(f"Error sending meeting deletion notification: {str(e)}")


# class MeetingCopyView(LoginRequiredMixin, View):
#     def post(self, request, pk, *args, **kwargs):
#         original = get_object_or_404(MeetingSchedule, pk=pk)
#         copied = MeetingSchedule.objects.create(
#             meeting_user=request.user,
#             title=f"{original.title} (Copy)",
#             description=original.description,
#             date=original.date,
#             start_time=original.start_time,
#             end_time=original.end_time,
#             meeting_url=original.meeting_url,
#             password=original.password,
#             is_sms=original.is_sms,
#         )
#         return JsonResponse({'status': 'success', 'copied_id': copied.pk})

# # Recording Views
# # class RecordingListView(LoginRequiredMixin, ListView):
# #     model = ClassRecording
# #     template_name = "recordings/recording_list.html"
# #     context_object_name = "recordings"

# #     def get_queryset(self):
# #         return ClassRecording.objects.select_related('meeting').order_by('-meeting__date')

# #     def get_context_data(self, **kwargs):
# #         context = super().get_context_data(**kwargs)
# #         # Total balance logic
# #         if self.request.user.is_superuser:
# #             total_investment = (
# #                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
# #             )
# #             context["total_balance"] = total_investment
# #         else:
# #             context["total_balance"] = self.request.user.balance
# #         return context

# class RecordingListView(LoginRequiredMixin, ListView):
#     model = ClassRecording
#     template_name = "recordings/recording_list.html"
#     context_object_name = "recordings"
#     paginate_by= 20

#     def get_queryset(self):
#         queryset = ClassRecording.objects.select_related('meeting').order_by('-meeting__date')

#         # filters
#         title = self.request.GET.get('title', '').strip()
#         from_date = self.request.GET.get('from_date', '').strip()
#         to_date = self.request.GET.get('to_date', '').strip()

#         if title:
#             queryset = queryset.filter(meeting__title__icontains=title)

#         if from_date:
#             queryset = queryset.filter(meeting__date__gte=from_date)

#         if to_date:
#             queryset = queryset.filter(meeting__date__lte=to_date)

#         return queryset

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)

#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance

#         # Preserve filter values in template
#         context['title_filter'] = self.request.GET.get('title', '')
#         context['from_date_filter'] = self.request.GET.get('from_date', '')
#         context['to_date_filter'] = self.request.GET.get('to_date', '')

#         return context

# class RecordingDetailView(LoginRequiredMixin, DetailView):
#     model = ClassRecording
#     template_name = "recordings/recording_detail.html"
#     context_object_name = "recording"

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
#         return context


# class RecordingCreateView(LoginRequiredMixin, CreateView):
#     model = ClassRecording
#     form_class = RecordingForm
#     template_name = "recordings/recording_form.html"
#     success_url = reverse_lazy("recording-list")

#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
#         # Only show meetings created by the current user
#         form.fields["meeting"].queryset = MeetingSchedule.objects.filter(
#             meeting_user=self.request.user, is_expired=False
#         )
#         return form

#     def form_valid(self, form):
#         messages.success(self.request, "Recording created successfully!")
#         return super().form_valid(form)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
#         return context


# class RecordingUpdateView(LoginRequiredMixin, UpdateView):
#     model = ClassRecording
#     form_class = RecordingForm
#     template_name = "recordings/recording_form.html"
#     success_url = reverse_lazy("recording-list")

#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
#         # Only show meetings created by the current user
#         form.fields["meeting"].queryset = MeetingSchedule.objects.filter(
#             meeting_user=self.request.user
#         )
#         return form

#     def form_valid(self, form):
#         messages.success(self.request, "Recording updated successfully!")
#         return super().form_valid(form)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = (
#                 User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             )
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
#         return context


# class RecordingDeleteView(LoginRequiredMixin, DeleteView):
#     model = ClassRecording
#     template_name = "recordings/recording_confirm_delete.html"
#     success_url = reverse_lazy("recording-list")

#     def delete(self, request, *args, **kwargs):
#         messages.success(request, "Recording deleted successfully!")
#         return super().delete(request, *args, **kwargs)


# @login_required
# def meeting_upload(request):
#     """
#     View function to upload Excel file containing meeting data.
#     The file should have the following columns:
#     - title
#     - description
#     - date (YYYY-MM-DD)
#     - start_time
#     - end_time
#     - meeting_url
#     - password
#     - is_sms (optional, True/False)
#     """
#     template = "meetings/meeting_upload.html"

#     if request.method == "POST" and "file" in request.FILES:
#         file = request.FILES["file"]

#         # Check if the file is an Excel file
#         if file.name.endswith(".xlsx"):
#             try:
#                 wb = load_workbook(file)
#                 sheet = wb.active
#                 success_count = 0
#                 error_count = 0
#                 errors = []

#                 # Start from row 2 (assuming row 1 has headers)
#                 for row in range(2, sheet.max_row + 1):
#                     try:
#                         # Extract values from the sheet
#                         title = sheet.cell(row=row, column=1).value
#                         description = sheet.cell(row=row, column=2).value
#                         date_value = sheet.cell(row=row, column=3).value
#                         start_time = str(sheet.cell(row=row, column=4).value)
#                         end_time = str(sheet.cell(row=row, column=5).value)
#                         meeting_url = sheet.cell(row=row, column=6).value
#                         password = sheet.cell(row=row, column=7).value

#                         # Check for is_sms column, default to False if not present
#                         is_sms = False
#                         if sheet.max_column >= 8:
#                             is_sms_value = sheet.cell(row=row, column=8).value
#                             if isinstance(is_sms_value, bool):
#                                 is_sms = is_sms_value
#                             elif isinstance(is_sms_value, str):
#                                 is_sms = is_sms_value.lower() in ["true", "yes", "1"]

#                         # Skip empty rows
#                         if not title:
#                             continue

#                         # Check if meeting with same title and date exists
#                         existing_meetings = MeetingSchedule.objects.filter(
#                             meeting_user=request.user, title=title, date=date_value
#                         )

#                         # Delete existing meetings with same title and date if found
#                         if existing_meetings.exists():
#                             existing_meetings.delete()

#                         # Create new meeting
#                         MeetingSchedule.objects.create(
#                             meeting_user=request.user,
#                             title=title,
#                             description=description,
#                             date=date_value,
#                             start_time=start_time,
#                             end_time=end_time,
#                             meeting_url=meeting_url,
#                             password=password,
#                             is_sms=is_sms,
#                         )
#                         success_count += 1

#                     except Exception as e:
#                         error_count += 1
#                         errors.append(f"Error at row {row}: {str(e)}")
#                         continue

#                 # Prepare result message
#                 if success_count > 0:
#                     messages.success(
#                         request, f"{success_count} meetings uploaded successfully."
#                     )

#                 if error_count > 0:
#                     error_message = f"{error_count} meetings failed to upload. "
#                     if errors:
#                         error_message += f"First error: {errors[0]}"
#                     messages.error(request, error_message)

#                 return redirect("meeting-list")

#             except Exception as e:
#                 messages.error(request, f"Error processing file: {str(e)}")
#         else:
#             messages.error(request, "Please upload a valid Excel (.xlsx) file!")
#     context = {}
#     if request.user.is_superuser:
#         # Calculate total investment from all users
#         total_investment = (
#             User.objects.all().aggregate(Sum("balance"))["balance__sum"] or 0
#         )
#         context["total_balance"] = total_investment
#     else:
#         context["total_balance"] = request.user.balance
#     return render(request, template, context)

# @login_required
# def download_meeting_template(request):
#     """Provides a template Excel file for meeting uploads"""
#     from django.http import HttpResponse
#     from openpyxl import Workbook

#     wb = Workbook()
#     ws = wb.active

#     # Add headers
#     headers = [
#         "Title",
#         "Description",
#         "Date (YYYY-MM-DD)",
#         "Start Time",
#         "End Time",
#         "Meeting URL",
#         "Password",
#         "Send SMS (True/False)",
#     ]

#     for col_num, header in enumerate(headers, 1):
#         ws.cell(row=1, column=col_num).value = header

#     # Example row
#     example_data = [
#         "Weekly Team Meeting",
#         "Regular team sync-up meeting",
#         "2025-05-27",
#         "10:00",
#         "11:00",
#         "https://example.com/meeting",
#         "pass123",
#         "False",
#     ]

#     for col_num, value in enumerate(example_data, 1):
#         ws.cell(row=2, column=col_num).value = value

#     # Create the HTTP response
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = (
#         'attachment; filename="meeting_upload_template.xlsx"'
#     )

#     wb.save(response)
#     return response

# class CalendarView(LoginRequiredMixin, View):
#     template_name = "meetings/calendar.html"

#     def get(self, request, *args, **kwargs):
#         today = timezone.localdate()
#         meetings = MeetingSchedule.objects.filter(date__gte=today).order_by("date", "start_time")

#         event_data = []
#         for meeting in meetings:
#             try:
#                 if not meeting.start_time or not meeting.end_time:
#                     continue

#                 start_datetime = datetime.combine(meeting.date, meeting.start_time)
#                 end_datetime = datetime.combine(meeting.date, meeting.end_time)

#                 if timezone.is_naive(start_datetime):
#                     start_datetime = timezone.make_aware(start_datetime)
#                 if timezone.is_naive(end_datetime):
#                     end_datetime = timezone.make_aware(end_datetime)

#                 # Handle crossing midnight
#                 if end_datetime < start_datetime:
#                     end_datetime += timedelta(days=1)

#                 event_data.append(
#                     {
#                         "id": meeting.id,
#                         "title": meeting.title,
#                         "start": start_datetime.isoformat(),
#                         "end": end_datetime.isoformat(),
#                     }
#                 )
#             except Exception as e:
#                 print(f"Error processing meeting id {meeting.id}: {e}")
#                 continue

#         context = {"event_data": json.dumps(event_data)}

#         if request.user.is_superuser:
#             total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = request.user.balance

#         return render(request, self.template_name, context)
