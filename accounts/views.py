import datetime
import os
from django.views.generic import CreateView, UpdateView
from django.utils import timezone
from django.http import HttpResponseRedirect
from .models import Property, User, PropertyContribution

import json
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Sum
from django.contrib import messages
from django.forms import modelformset_factory
from django.shortcuts import redirect
from django.contrib import messages
from .forms import BeneficiaryForm
from .models import Beneficiary
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction

from django.views.generic import ListView
from decimal import Decimal, ROUND_HALF_UP
from .models import Expense, User, ExpensePayment
from .forms import OfficeExpensePaymentForm
from decimal import Decimal, InvalidOperation
  
from datetime import datetime, date
from django.db.models import Sum, Count 
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from .forms import ExpenseForm, ExpensePaymentForm
from .models import Expense, ExpensePayment, DeductionHistory
import paypalrestsdk
from base64 import b64encode
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
import weasyprint
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import LoginView
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from email.mime.image import MIMEImage
from django.db.models import Prefetch, Sum, Case, When, IntegerField
from django.db.models.functions import Coalesce
from django.http import FileResponse, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy

from django.utils.html import strip_tags
from django.utils.timezone import now
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.http import require_GET
from django.utils.dateparse import parse_date
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.utils.decorators import method_decorator
import uuid
from square.client import Client
from accounts.utils import (
    calculate_user_investment_summary,
    generate_stock_certificate_image,
)
from hfallmedia.models import HeroVideo

from .forms import (
    AnnouncementForm,
    PaymentApprovalForm,
    PaymentForm,
    PropertyForm,
    PropertyImageFormSet,
    UserLoginForm,
    UserRegistrationForm,
    UserUpdateForm,
    HelpForm
)
from .models import (
    Agreement,
    Bank,
    Payment,
    Property,
    PropertyImage,
    Story,
    User,
    UserAgreement
)
import stripe
from django.http import  JsonResponse
from django.urls import reverse

stripe.api_key = settings.STRIPE_SECRET_KEY

def home(request):
    properties = Property.objects.filter(status="ready_to_sell").order_by(
        "-listed_date"
    )[:10]
    video = HeroVideo.objects.order_by("-created_at").first()
    return render(request, "home.html", {"properties": properties, "video": video})


def property_images(request, property_id):
    images = PropertyImage.objects.filter(property_id=property_id)
    data = [{"image": img.image.url} for img in images]
    return JsonResponse(data, safe=False)


def register(request):
    active_agreement = (
        Agreement.objects.filter(is_active=True).order_by("-created_at").first()
    )

    if request.method == "POST":
        form = UserRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            raw_password = form.cleaned_data.get("password")
            user = form.save()
            if active_agreement:
                UserAgreement.objects.create(
                    user=user,
                    agreement=active_agreement,
                )

            context = {
                "user_full_name": user.get_full_name(),
                "user": user,
                "password": raw_password,
            }
            html_content = render_to_string("emails/welcome_email.html", context)
            text_content = strip_tags(html_content)
            subject = "Welcome to HFall Realty!"
            from_email = settings.DEFAULT_FROM_EMAIL
            to = user.email

            email = EmailMultiAlternatives(subject, text_content, from_email, [to])
            email.attach_alternative(html_content, "text/html")
            email.send()

            messages.success(
                request, "Registration successful. A confirmation email has been sent."
            )
            return redirect("accounts:success-message")
    else:
        form = UserRegistrationForm()
    return render(
        request, "register.html", {"form": form, "agreement": active_agreement}
    )


def success_message(request):
    user = request.user
    context = {"user": user}
    return render(request, "snippet/successMessage.html", context)


class CustomLoginView(LoginView):
    template_name = "login.html"
    authentication_form = UserLoginForm


def get_property_stories(request, property_id):
    stories = Story.objects.filter(related_property_id=property_id).order_by(
        "-created_at"
    )
    data = [
        {
            "message": story.message,
            "timestamp": story.created_at.strftime("%b %d, %Y %I:%M %p"),
        }
        for story in stories
    ]
    return JsonResponse({"stories": data})


@login_required
def download_stock_certificate(request):
    # You can customize how these values are fetched
    if request.user.is_authenticated:
        name = request.user.get_full_name() or request.user.first_name
    else:
        name = "Anonymous Holder"

    shares = 1

    # Generate the PDF using the utility
    pdf_buffer = generate_stock_certificate_image(name=name, shares=shares)

    # Return the response for downloading
    return FileResponse(
        pdf_buffer,
        as_attachment=False,  # Let the browser display it
        filename="certificate.pdf",
        content_type='application/pdf'
    )


@login_required
def dashboard(request):
    summary = calculate_user_investment_summary(request.user)
    try:
        properties_bought = Property.objects.filter(status="bought")
        properties_wishlist = Property.objects.filter(status="wishlist")
        properties_sold = Property.objects.filter(status="sold")
    except properties_bought.DoesNotExist:
        properties_bought = None
    except properties_wishlist.DoesNotExist:
        properties_wishlist = None
    except properties_sold.DoesNotExist:
        properties_sold = None

    active_users = User.objects.filter(is_active=True, investor=True)
    # Calculate company-wide metrics
    total_balance = active_users.aggregate(Sum("balance"))["balance__sum"] or Decimal(
        "0.00"
    )

    # Get properties that have been purchased
    purchased_properties = Property.objects.filter(
        status__in=["bought", "ready_to_sell", "pending", "sold"]
    )

    # Calculate total invested (auction price)
    total_invested = purchased_properties.aggregate(
        total=Coalesce(Sum("buying_price"), Decimal("0.00"))
    )["total"]

    # Calculate total repair costs (service cost)
    total_repair_cost = purchased_properties.aggregate(
        total=Coalesce(Sum("service_cost"), Decimal("0.00"))
    )["total"]

    # Calculate total expenses not associated with properties
    additional_expenses = Expense.objects.filter(
        status="approved", property__isnull=True
    ).aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))["total"]

    # Initial total balance (before investments)
    initial_total_balance = (
        total_balance + total_invested + total_repair_cost + additional_expenses
    )

    # Calculate remaining balance
    remaining_balance = total_balance

    # Get current user's metrics
    user = request.user
    user_balance = user.balance

    # Calculate user's contribution percentage to the total current balance
    contribution_percentage = 0
    if total_balance > Decimal("0.00"):
        contribution_percentage = (user_balance / total_balance) * Decimal("100")
    if total_balance > Decimal("0.00"):
        user_already_invested = (total_invested + total_repair_cost) * (
            user_balance / total_balance
        )
    else:
        user_already_invested = Decimal("0.00")

    user_total_investment = user_balance + user_already_invested

    user_contribution = user_already_invested
    sold_properties = Property.objects.filter(status="sold")
    total_profit = Decimal("0.00")
    for prop in sold_properties:
        if prop.selling_price and prop.buying_price:
            property_profit = prop.selling_price - (
                prop.buying_price + (prop.service_cost or Decimal("0.00"))
            )
            total_profit += property_profit
    user_profit = (
        total_profit * (contribution_percentage / Decimal("100"))
        if total_profit > Decimal("0.00")
        else Decimal("0.00")
    )
    properties_data = []
    for i, prop in enumerate(
        purchased_properties, 0
    ):  # Limit to 2 properties as in the image
        if not prop.buying_price:
            continue

        total_cost = prop.buying_price + (prop.service_cost or Decimal("0.00"))

        profit_loss = (prop.selling_price or Decimal("0.00")) - total_cost

        property_data = {
            "id": prop.id,
            "title": prop.property_name,
            "number": i + 1,
            "bought": prop.buying_price,
            "repair_cost": prop.service_cost or Decimal("0.00"),
            "sold": prop.selling_price or Decimal("0.00"),
            "total": profit_loss,
        }
        properties_data.append(property_data)

    context = {
        "properties_bought": properties_bought,
        "properties_wishlist": properties_wishlist,
        "properties_sold": properties_sold,
        # Company-wide data
        "company_total_balance": initial_total_balance,
        "company_invested": total_invested,
        "company_repair_cost": total_repair_cost,
        "company_remaining": remaining_balance,
        # Current user data
        "user_total_investment": user_total_investment,
        "user_contribution": user_already_invested,  # This shows how much they've invested in properties
        "user_remaining": user_balance,  # This shows their remaining balance
        "user_contribution_percentage": round(contribution_percentage, 1),
        "user_profit": user_profit,
        # Properties data
        "properties_data": properties_data,
        # summary data
        "summary": summary,
    }

    if request.user.is_superuser:
        # Calculate total investment from all users
        total_investment = (
            User.objects.all().aggregate(Sum("balance"))["balance__sum"] or 0
        )
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance

    return render(request, "dashboard.html", context)


@login_required
def user_logout(request):
    """
    Log out the user and redirect to login page
    """
    logout(request)
    messages.success(request, "You have been successfully logged out.")
    return redirect("accounts:login")



@login_required
@xframe_options_exempt
def profile_view(request):
    approved_payments = Payment.objects.filter(
        user=request.user, status="approved"
    ).order_by("-approved_at")

    # Get all uploaded agreements from all users
    all_uploaded_agreements = (
        UserAgreement.objects.filter(uploaded_file__isnull=False)
        .exclude(uploaded_file="")
        .order_by("-created_at")
    )

    context = {
        "user": request.user,
        "approved_payments": approved_payments,
        "all_uploaded_agreements": all_uploaded_agreements,
    }

    if request.GET.get("preview") == "pdf" or request.GET.get("download") == "pdf":
        image_context = {}

        # Handle user personal image
        if request.user.personal_image:
            try:
                image_path = request.user.personal_image.path
                if os.path.exists(image_path):
                    with open(image_path, "rb") as image_file:
                        encoded_string = b64encode(image_file.read()).decode()
                        image_context["user_image_data"] = (
                            f"data:image/{os.path.splitext(image_path)[1][1:].lower()};base64,{encoded_string}"
                        )
            except Exception as e:
                print(f"Error processing personal image: {e}")

        # Handle user photo ID
        if request.user.photo_id:
            try:
                photo_id_path = request.user.photo_id.path
                if os.path.exists(photo_id_path):
                    with open(photo_id_path, "rb") as photo_id_file:
                        encoded_string = b64encode(photo_id_file.read()).decode()
                        image_context["photo_id_data"] = (
                            f"data:image/{os.path.splitext(photo_id_path)[1][1:].lower()};base64,{encoded_string}"
                        )
            except Exception as e:
                print(f"Error processing photo ID: {e}")

        # Handle logo image
        try:
            logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/logo.png")
            footer_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf_footer.png")

            # If STATIC_ROOT doesn't exist or file not found, try STATICFILES_DIRS
            if not (os.path.exists(logo_path) and os.path.exists(footer_path)):
                for static_dir in settings.STATICFILES_DIRS:
                    test_logo = os.path.join(static_dir, "assets/images/logo.png")
                    test_footer = os.path.join(static_dir, "assets/images/pdf_footer.png")
                    if os.path.exists(test_logo) and os.path.exists(test_footer):
                        logo_path = test_logo
                        footer_path = test_footer
                        break

            # Process logo image
            if os.path.exists(logo_path):
                with open(logo_path, "rb") as logo_file:
                    encoded_logo = b64encode(logo_file.read()).decode()
                    image_context["logo_data"] = f"data:image/png;base64,{encoded_logo}"

            # Process footer image
            if os.path.exists(footer_path):
                with open(footer_path, "rb") as footer_file:
                    encoded_footer = b64encode(footer_file.read()).decode()
                    image_context["footer_data"] = f"data:image/png;base64,{encoded_footer}"

        except Exception as e:
            print(f"Error processing logo/footer: {e}")

        image_context["base_url"] = request.build_absolute_uri("/")
        context.update(image_context)

        # Render HTML template
        html = render_to_string("profile_pdf.html", context)

        # Create PDF response
        response = HttpResponse(content_type="application/pdf")

        # Set content disposition based on whether it's a preview or download
        if request.GET.get("download") == "pdf":
            response["Content-Disposition"] = 'attachment; filename="profile.pdf"'
        else:
            response["Content-Disposition"] = 'inline; filename="profile.pdf"'

        # Generate PDF with WeasyPrint
        pdf = weasyprint.HTML(
            string=html, base_url=request.build_absolute_uri("/")
        ).write_pdf()
        response.write(pdf)
        return response

    if request.user.is_superuser:
        # Calculate total investment from all users
        total_investment = (
            User.objects.all().aggregate(Sum("balance"))["balance__sum"] or 0
        )
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance

    # Normal view
    return render(request, "profile.html", context)


class UserProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = "profile_update.html"
    success_url = reverse_lazy("accounts:profile")

    def get_object(self):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_superuser:
            total_investment = (
                User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            )
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance
        return context
    def form_valid(self, form):
        
        names = self.request.POST.getlist('beneficiary_name[]')
        percentages = self.request.POST.getlist('beneficiary_percentage[]')
        
        beneficiaries = []
        for name, percentage in zip(names, percentages):
           
            if name.strip() and percentage.strip():
                try:
                    beneficiaries.append({
                        'name': name.strip(),
                        'percentage': float(percentage)
                    })
                except (ValueError, TypeError):
                    pass 
        
        form.instance.beneficiaries = beneficiaries
        messages.success(self.request, "Profile updated successfully!")
        return super().form_valid(form)

class  member_list(LoginRequiredMixin, ListView):
    model = User
    template_name = "member_list.html"
    context_object_name = "members"
    paginate_by = 20

    def get_queryset(self):
        return User.objects.all().order_by("id")  

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance

        return context


@login_required
def member_list_card(request):
    members = User.objects.all()
    for i in members:
        print(i.personal_image)
    context = {"members": members}
    if request.user.is_superuser:
        # Calculate total investment from all users
        total_investment = (
            User.objects.all().aggregate(Sum("balance"))["balance__sum"] or 0
        )
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "member_list_card.html", context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
@xframe_options_exempt
def member_detail(request, pk):
    member = get_object_or_404(User, pk=pk)

    approved_payments = Payment.objects.filter(user=member, status="approved").order_by("-approved_at")

    # All uploaded agreements
    all_uploaded_agreements = (
        UserAgreement.objects.filter(uploaded_file__isnull=False)
        .exclude(uploaded_file="")
        .order_by("-created_at")
    )

    context = {
        "member": member,
        "approved_payments": approved_payments,
        "all_uploaded_agreements": all_uploaded_agreements,
    }

    if request.GET.get("preview") == "pdf" or request.GET.get("download") == "pdf":
        image_context = {}

        # Personal Image
        if member.personal_image:
            try:
                image_path = member.personal_image.path
                if os.path.exists(image_path):
                    with open(image_path, "rb") as image_file:
                        encoded_string = b64encode(image_file.read()).decode()
                        image_context["user_image_data"] = f"data:image/{os.path.splitext(image_path)[1][1:].lower()};base64,{encoded_string}"
            except Exception as e:
                print(f"Error processing personal image: {e}")

        # Photo ID
        if member.photo_id:
            try:
                photo_id_path = member.photo_id.path
                if os.path.exists(photo_id_path):
                    with open(photo_id_path, "rb") as photo_id_file:
                        encoded_string = b64encode(photo_id_file.read()).decode()
                        image_context["photo_id_data"] = f"data:image/{os.path.splitext(photo_id_path)[1][1:].lower()};base64,{encoded_string}"
            except Exception as e:
                print(f"Error processing photo ID: {e}")

        # Logo and Footer
        try:
            logo_path = os.path.join(settings.STATIC_ROOT, "assets/images/logo.png")
            footer_path = os.path.join(settings.STATIC_ROOT, "assets/images/pdf_footer.png")

            if not (os.path.exists(logo_path) and os.path.exists(footer_path)):
                for static_dir in settings.STATICFILES_DIRS:
                    test_logo = os.path.join(static_dir, "assets/images/logo.png")
                    test_footer = os.path.join(static_dir, "assets/images/pdf_footer.png")
                    if os.path.exists(test_logo) and os.path.exists(test_footer):
                        logo_path = test_logo
                        footer_path = test_footer
                        break

            if os.path.exists(logo_path):
                with open(logo_path, "rb") as logo_file:
                    encoded_logo = b64encode(logo_file.read()).decode()
                    image_context["logo_data"] = f"data:image/png;base64,{encoded_logo}"

            if os.path.exists(footer_path):
                with open(footer_path, "rb") as footer_file:
                    encoded_footer = b64encode(footer_file.read()).decode()
                    image_context["footer_data"] = f"data:image/png;base64,{encoded_footer}"

        except Exception as e:
            print(f"Error processing logo/footer: {e}")

        image_context["base_url"] = request.build_absolute_uri("/")
        context.update(image_context)

        # Render PDF template
        html = render_to_string("member_pdf.html", context)
        response = HttpResponse(content_type="application/pdf")

        if request.GET.get("download") == "pdf":
            response["Content-Disposition"] = 'attachment; filename="member_profile.pdf"'
        else:
            response["Content-Disposition"] = 'inline; filename="member_profile.pdf"'

        pdf = weasyprint.HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()
        response.write(pdf)
        return response
    if request.user.is_superuser:
        # Calculate total investment from all users
        total_investment = (
            User.objects.all().aggregate(Sum("balance"))["balance__sum"] or 0
        )
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "member_detail.html", context)


class PropertyListView(LoginRequiredMixin, ListView):
    model = Property
    template_name = "property_list.html"
    context_object_name = "properties"

    def get_queryset(self):
        # Prefetch stories ordered by latest
        story_prefetch = Prefetch(
            "stories",
            queryset=Story.objects.only(
                "id", "message", "created_at", "related_property"
            ).order_by("-created_at"),
        )

        queryset = Property.objects.prefetch_related("images", story_prefetch).order_by(
            "-created_at"
        )

        status = self.request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["status_choices"] = Property.STATUS_CHOICES
        context["selected_status"] = self.request.GET.get("status", "")
        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = (
                User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            )
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance
        return context


class PropertyDetailView(DetailView):
    model = Property
    template_name = "property_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["property_images"] = self.object.images.all()
        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = (
                User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            )
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance
        return context


class SuperuserRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser


class PropertyUserRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and (
            self.request.user.is_superuser
            or getattr(self.request.user, "is_property", False)
        )


class PropertyCreateView(PropertyUserRequiredMixin, CreateView):
    model = Property
    form_class = PropertyForm
    template_name = "property_form.html"
    success_url = reverse_lazy("accounts:property_list")
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["image_formset"] = PropertyImageFormSet(
            self.request.POST or None,
            self.request.FILES or None
        )
        all_users = User.objects.filter(is_active=True, investor=True)
        context['all_users'] = all_users
        context['user_contributions'] = {}  # Empty for create - template expects this key
        context["total_balance"] = (
            User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            if self.request.user.is_superuser
            else self.request.user.balance
        )
        # Default investment date
        context["default_investment_date"] = date.today().strftime('%Y-%m-%d')
        
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        image_formset = context["image_formset"]
        
        print("\n" + "="*80)
        print("🆕 CREATING NEW PROPERTY WITH MULTI-INVESTMENT SUPPORT")
        print("="*80)
        
        with transaction.atomic():
            form.instance.listed_by = self.request.user
            buying_price = form.cleaned_data.get('buying_price') or 0
            service_cost = form.cleaned_data.get('service_cost') or 0
            form.instance.acquisition_cost = buying_price + service_cost if (buying_price or service_cost) else None
            
            # Save property first
            self.object = form.save(commit=False)
            self.object.save()
            
            print(f"✅ Property saved: {self.object.property_name}")
            print(f"   Status: {self.object.status}")
            print(f"   Buying Price: ${self.object.buying_price}")
            print(f"   Service Cost: ${self.object.service_cost}")
            
            if self.object.status == 'bought':
                # Parse multi-investment data from POST
                investments_list = []
                investment_dates_list = []
                selected_contributors = set()
                
                print("\n📊 Parsing Investment Data...")
                
                # Parse all invest_ fields with sequence numbers
                for key in self.request.POST:
                    if key.startswith('invest_'):
                        # Format: invest_{user_id}_{sequence}
                        parts = key.replace('invest_', '').split('_')
                        
                        if len(parts) >= 2:
                            user_id = int(parts[0])
                            sequence = int(parts[1])
                        else:
                            # Old format fallback: invest_{user_id}
                            user_id = int(parts[0])
                            sequence = 1
                        
                        amount_str = self.request.POST.get(key, '').strip()
                        
                        try:
                            amount = Decimal(amount_str)
                            
                            # Check if checkbox is selected
                            checkbox_key = f'select_user_{user_id}_{sequence}'
                            is_selected = self.request.POST.get(checkbox_key) is not None
                            
                            if amount > 0 and is_selected:
                                # Check if fixed
                                fixed_key = f'fixed_{user_id}_{sequence}'
                                is_fixed = self.request.POST.get(fixed_key) is not None
                                
                                investments_list.append({
                                    'user_id': user_id,
                                    'invest_amount': amount,
                                    'is_fixed': is_fixed,
                                    'sequence': sequence
                                })
                                
                                selected_contributors.add(user_id)
                                
                                # Get investment date
                                date_key = f'date_{user_id}_{sequence}'
                                date_str = self.request.POST.get(date_key, '').strip()
                                
                                if date_str:
                                    try:
                                        inv_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                    except ValueError:
                                        inv_date = self.object.buying_date or date.today()
                                else:
                                    inv_date = self.object.buying_date or date.today()
                                
                                investment_dates_list.append({
                                    'user_id': user_id,
                                    'sequence': sequence,
                                    'date': inv_date
                                })
                                
                                print(f"   ✓ User {user_id} Investment #{sequence}: ${amount} (Fixed: {is_fixed}, Date: {inv_date})")
                        
                        except (ValueError, InvalidOperation, AttributeError) as e:
                            print(f"   ✗ Error parsing {key}: {e}")
                            continue
                
                print(f"\n📝 Total Investments Collected: {len(investments_list)}")
                print(f"   Unique Contributors: {len(selected_contributors)}")
                
                if investments_list:
                    # Set contributors
                    self.object.contributors.set(
                        User.objects.filter(id__in=selected_contributors)
                    )
                    
                    print("\n💰 Processing Investments...")
                    
                    # Use the new multi-investment method
                    success = self.object.deduct_property_costs_with_multiple_investments(
                        investments_list, 
                        investment_dates_list
                    )
                    
                    if not success:
                        print("❌ Investment processing failed!")
                        form.add_error(None, "Invalid investment data or insufficient balance.")
                        return self.form_invalid(form)
                    
                    print("✅ All investments processed successfully!")
                else:
                    print("⚠️  No valid investments found")
            
            # Save images
            if image_formset.is_valid():
                image_formset.instance = self.object
                image_formset.save()
                print("✅ Images saved")
        
        print("="*80)
        print("✅ PROPERTY CREATION COMPLETED SUCCESSFULLY")
        print("="*80 + "\n")
        
        return HttpResponseRedirect(self.get_success_url())

class PropertyUpdateView(PropertyUserRequiredMixin, UpdateView):
    model = Property
    form_class = PropertyForm
    template_name = "property_form.html"
    success_url = reverse_lazy("accounts:property_list")
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # ✅ Get fresh user data - force evaluation with list()
        all_users = list(
            User.objects.filter(
                is_active=True, 
                investor=True
            ).order_by('id')
        )
        
        # Build user contributions map
        user_contributions = {}
        for user in all_users:
            # Get all contributions for this user, force fresh query
            contributions = list(
                PropertyContribution.objects.filter(
                    property=self.object, 
                    user=user
                ).order_by('investment_sequence')
            )
            
            if contributions:
                user_contributions[user.id] = contributions
        
        # Build image formset
        context["image_formset"] = PropertyImageFormSet(
            self.request.POST or None,
            self.request.FILES or None,
            instance=self.object
        )
        
        context["all_users"] = all_users
        context["user_contributions"] = user_contributions
        
        # Calculate total balance from fresh data
        if self.request.user.is_superuser:
            context["total_balance"] = sum(user.balance for user in all_users)
        else:
            # Get fresh balance for current user
            current_user = User.objects.get(pk=self.request.user.pk)
            context["total_balance"] = current_user.balance
        
        # Set default investment date
        context["default_investment_date"] = (
            self.object.buying_date.strftime('%Y-%m-%d') 
            if self.object.buying_date 
            else date.today().strftime('%Y-%m-%d')
        )
        
        return context
    
    def form_valid(self, form):
        old_property = Property.objects.get(pk=self.object.pk)
        old_status = old_property.status
        
        print("\n" + "="*80)
        print(f"🔄 UPDATING PROPERTY: {old_property.property_name}")
        print("="*80)
        
        buying_price = form.cleaned_data.get("buying_price") or 0
        service_cost = form.cleaned_data.get("service_cost") or 0
        form.instance.acquisition_cost = (
            buying_price + service_cost if (buying_price or service_cost) else None
        )
        
        self.object = form.save(commit=False)
        self.object.save()
        
        if self.object.status == "bought":
            # Get all existing contributions with their sequences
            existing_contributions = PropertyContribution.objects.filter(
                property=self.object
            ).select_related('user')
            
            # Create mapping: {user_id: {sequence: contribution_object}}
            existing_map = {}
            for contrib in existing_contributions:
                if contrib.user_id not in existing_map:
                    existing_map[contrib.user_id] = {}
                existing_map[contrib.user_id][contrib.investment_sequence] = contrib
            
            investments_list = []
            investment_dates_list = []
            selected_contributors = set()
            
            # Track which existing contributions are being submitted
            submitted_contributions = set()  # {(user_id, sequence)}
            
            print("\n📊 Parsing Updated Investment Data...")
            
            # Parse all invest_ fields with sequence
            for key in self.request.POST:
                if key.startswith("invest_"):
                    parts = key.replace("invest_", "").split("_")
                    if len(parts) >= 2:
                        user_id = int(parts[0])
                        sequence = int(parts[1])
                    else:
                        user_id = int(parts[0])
                        sequence = 1
                    
                    amount_str = self.request.POST.get(key, '').strip()
                    try:
                        amount = Decimal(amount_str)
                        
                        # Check if checkbox is selected
                        checkbox_key = f'select_user_{user_id}_{sequence}'
                        is_selected = self.request.POST.get(checkbox_key) is not None
                        
                        if amount > 0 and is_selected:
                            fixed_key = f'fixed_{user_id}_{sequence}'
                            is_fixed = self.request.POST.get(fixed_key) is not None
                            
                            investments_list.append({
                                'user_id': user_id,
                                'invest_amount': amount,
                                'is_fixed': is_fixed,
                                'sequence': sequence
                            })
                            selected_contributors.add(user_id)
                            submitted_contributions.add((user_id, sequence))
                            
                            # Get investment date
                            date_key = f'date_{user_id}_{sequence}'
                            date_str = self.request.POST.get(date_key, '').strip()
                            if date_str:
                                try:
                                    inv_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                                except ValueError:
                                    inv_date = self.object.buying_date or date.today()
                            else:
                                # Use existing date if available
                                if user_id in existing_map and sequence in existing_map[user_id]:
                                    inv_date = existing_map[user_id][sequence].investment_date or self.object.buying_date or date.today()
                                else:
                                    inv_date = self.object.buying_date or date.today()
                            
                            investment_dates_list.append({
                                'user_id': user_id,
                                'sequence': sequence,
                                'date': inv_date
                            })
                            
                            print(f"   ✓ User {user_id} Investment #{sequence}: ${amount} (Fixed: {is_fixed})")
                    
                    except (ValueError, InvalidOperation, AttributeError) as e:
                        print(f"   ✗ Error parsing {key}: {e}")
                        continue
            
            print(f"\n📝 Total Investments: {len(investments_list)}")
            print(f"   Selected Contributors: {selected_contributors}")
            
            # ✅ FIX: Check for NEW contributors OR NEW sequences
            existing_user_ids = set(existing_map.keys())
            truly_new_user_ids = selected_contributors - existing_user_ids
            
            # Check for new sequences from existing users
            new_sequences_added = []
            for user_id, sequence in submitted_contributions:
                if user_id in existing_map:
                    if sequence not in existing_map[user_id]:
                        new_sequences_added.append((user_id, sequence))
            
            # ⚠️ CRITICAL FIX: Check if any contributions were REMOVED
            removed_contributions = []
            for user_id, sequences in existing_map.items():
                for sequence in sequences:
                    if (user_id, sequence) not in submitted_contributions:
                        removed_contributions.append((user_id, sequence))
            
            should_refund = bool(truly_new_user_ids or new_sequences_added or removed_contributions)
            
            if should_refund:
                if truly_new_user_ids:
                    print(f"\n🆕 NEW CONTRIBUTORS: {truly_new_user_ids}")
                if new_sequences_added:
                    print(f"\n➕ NEW SEQUENCES ADDED:")
                    for user_id, seq in new_sequences_added:
                        print(f"   User {user_id} Sequence #{seq}")
                if removed_contributions:
                    print(f"\n🗑️  CONTRIBUTIONS REMOVED:")
                    for user_id, seq in removed_contributions:
                        print(f"   User {user_id} Sequence #{seq}")
                
                print("\n🔄 Refunding all existing contributions...")
                self.object.refund_all_contributions()
                print("   ✅ Refund completed")
            else:
                print("\n✅ No structural changes - updating existing contributions only")
            
            if investments_list:
                # Set contributors
                self.object.contributors.set(
                    User.objects.filter(id__in=selected_contributors)
                )
                
                print("\n💰 Processing Updated Investments...")
                success = self.object.deduct_property_costs_with_multiple_investments(
                    investments_list, 
                    investment_dates_list
                )
                
                if not success:
                    print("❌ Update failed!")
                    form.add_error(None, "Invalid investment or insufficient balance.")
                    return self.form_invalid(form)
                
                print("✅ Update successful!")
        
        elif old_status == "bought" and self.object.status == "sold":
            print("\n💵 Property sold - distributing proceeds...")
            self.object.distribute_sale_proceeds()
        
        elif old_status == "bought" and self.object.status not in ["bought", "sold"]:
            print("\n↩️  Status changed - refunding contributions...")
            self.object.refund_all_contributions()
        
        # Handle image formset
        context = self.get_context_data()
        image_formset = context["image_formset"]
        if image_formset.is_valid():
            image_formset.save()
        
        print("="*80 + "\n")
        return HttpResponseRedirect(self.get_success_url())
class PropertyGalleryView(DetailView):
    model = Property
    template_name = "property_gallery.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["property_images"] = self.object.images.all()
        return context


@require_GET
def property_api_detail(request, pk):
    """API view to get property details in JSON format for the modal"""
    property = get_object_or_404(Property, pk=pk)

    # Get primary image or first image
    primary_image = property.images.filter(is_primary=True).first()
    if not primary_image:
        primary_image = property.images.first()

    # Get all property images
    all_images = []
    for img in property.images.all():
        all_images.append({"url": img.image.url, "is_primary": img.is_primary})

    # Format property data for JSON response
    data = {
        "id": property.id,
        "title": property.property_name,
        "description": property.description,
        "address": property.address,
        "city": property.city,
        "state": property.state,
        "zip": property.zip_code,
        "status": property.status,
        "bedrooms": property.bedrooms,
        "bathrooms": property.bathrooms,
        "dining_rooms": property.dining_rooms,
        "squareFeet": property.square_feet,
        "price": float(property.final_price),
        "buying_price": float(property.buying_price),
        "service_cost": float(property.service_cost),
        "profit": float(property.profit),
        "mainImage": primary_image.image.url if primary_image else None,
        "images": all_images,
        "listed_date": (
            property.listed_date.strftime("%B %d, %Y") if property.listed_date else None
        ),
    }

    return JsonResponse(data)


class PropertyDeleteView(PropertyUserRequiredMixin, DeleteView):
    model = Property
    template_name = "property_confirm_delete.html"
    success_url = reverse_lazy("accounts:property_list")


@login_required
def payment_banks(request):
    banks = Bank.objects.filter(is_active=True, is_card=False)

    context = {"banks": banks}
    # Total balance logic
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "payment_banks.html", context)


@login_required
def make_payment(request, bank_id):
    bank = get_object_or_404(Bank, id=bank_id, is_active=True)

    if request.method == "POST":
        form = PaymentForm(request.POST, request.FILES)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.user = request.user
            payment.bank = bank
            payment.save()
            user= payment.user

            # ✅ Send email to admin
            subject = f"New Payment Submission (Pending Review) by {request.user.get_full_name()}"
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = [user.email]
            context = {
                "user": request.user,
                "bank": bank,
                "payment": payment,
                "date": now().strftime("%d %b %Y"),
            }

            html_content = render_to_string("emails/admin_payment_notification.html", context)
            text_content = f"""
            A new payment has been submitted by {request.user.get_full_name()} on {context['date']}.
            Bank: {bank.name}
            Amount: {payment.amount}
            Please log in to review it.

            Regards,
            HFall Payment System
            """

            email = EmailMultiAlternatives(subject, text_content.strip(), from_email, to_email)
            email.attach_alternative(html_content, "text/html")
            email.send()

            messages.success(
                request,
                "Payment submitted successfully. It will be reviewed by an administrator.",
            )
            return redirect("accounts:my_payments")
    else:
        form = PaymentForm()

    context = {"form": form, "bank": bank}
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance

    return render(request, "make_payment.html", context)



@method_decorator(login_required, name="dispatch")
class my_payments(ListView):
    model = Payment
    template_name = "my_payments.html"
    context_object_name = "payments"
    paginate_by = 10000

    def get_queryset(self):
        payments = Payment.objects.filter(user=self.request.user).order_by("-created_at")

        # Filters
        from_date = self.request.GET.get("from_date")
        to_date = self.request.GET.get("to_date")
        status = self.request.GET.get('status', 'approved') 
        bank_id = self.request.GET.get("bank")

        if from_date:
            payments = payments.filter(created_at__date__gte=parse_date(from_date))
        if to_date:
            payments = payments.filter(created_at__date__lte=parse_date(to_date))
        if status and status != "all":
            payments = payments.filter(status=status)
        if bank_id and bank_id != "all":
            payments = payments.filter(bank_id=bank_id)

        return payments

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        payments = self.get_queryset()

        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            context_total_balance = total_investment
        else:
            context_total_balance = self.request.user.balance

        # Check if filters applied
        from_date = self.request.GET.get("from_date")
        to_date = self.request.GET.get("to_date")
        status = self.request.GET.get("status")
        bank_id = self.request.GET.get("bank")

        is_filtered = bool(from_date or to_date or (status and status != "all") or (bank_id and bank_id != "all"))

        # Filter sum
        if is_filtered:
            filter_sum = payments.aggregate(Sum("amount"))["amount__sum"] or 0
            
        else:
            filter_sum = Payment.objects.filter(
                user=self.request.user,
                status="approved"   
            ).aggregate(Sum("amount"))["amount__sum"] or 0
        # else:
        #     filter_sum = Payment.objects.filter(user=self.request.user).aggregate(Sum("amount"))["amount__sum"] or 0

        context.update({
            "total_balance": context_total_balance,
            "filter_sum": filter_sum,
            "banks": Bank.objects.all().order_by("name"),
            "statuses": Payment.STATUS_CHOICES,
            "from_date": from_date or "",
            "to_date": to_date or "",
            "selected_status": status or "all",
            "selected_bank": int(bank_id) if bank_id and bank_id.isdigit() else "all",
        })

        return context

# @method_decorator([login_required, user_passes_test(lambda u: u.is_superuser)], name="dispatch")
# class pending_payments(ListView):
#     model = Payment
#     template_name = "pending_payments.html"
#     context_object_name = "payments"
#     paginate_by = 1000

#     def get_queryset(self):
#         payments = Payment.objects.filter(status="pending").order_by("-created_at")

#         # Filters
#         user_id = self.request.GET.get("user")
#         bank_id = self.request.GET.get("bank")
#         from_date = self.request.GET.get("from_date")
#         to_date = self.request.GET.get("to_date")

#         if user_id:
#             payments = payments.filter(user_id=user_id)
#         if bank_id:
#             payments = payments.filter(bank_id=bank_id)
#         if from_date:
#             payments = payments.filter(created_at__date__gte=parse_date(from_date))
#         if to_date:
#             payments = payments.filter(created_at__date__lte=parse_date(to_date))

#         return payments

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)

#         user_id = self.request.GET.get("user")
#         bank_id = self.request.GET.get("bank")
#         from_date = self.request.GET.get("from_date")
#         to_date = self.request.GET.get("to_date")

#         # Add filters to context
#         context.update({
#             "users": User.objects.all().order_by("short_name"),
#             "banks": Bank.objects.all().order_by("name"),
#             "selected_user": user_id,
#             "selected_bank": bank_id,
#             "from_date": from_date,
#             "to_date": to_date,
#         })

#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance

#         return context
@method_decorator(login_required, name="dispatch")
class pending_payments(ListView):
    model = Payment
    template_name = "pending_payments.html"
    context_object_name = "payments"
    paginate_by = 1000

    def get_queryset(self):
        user = self.request.user

        # Base queryset
        if user.is_superuser:
            payments = Payment.objects.filter(status="pending")
        else:
            payments = Payment.objects.filter(
                status="pending",
                user=user
            )

        payments = payments.order_by("-created_at")

        # Filters (superuser only)
        if user.is_superuser:
            user_id = self.request.GET.get("user")
            bank_id = self.request.GET.get("bank")
            from_date = self.request.GET.get("from_date")
            to_date = self.request.GET.get("to_date")

            if user_id:
                payments = payments.filter(user_id=user_id)
            if bank_id:
                payments = payments.filter(bank_id=bank_id)
            if from_date:
                payments = payments.filter(created_at__date__gte=parse_date(from_date))
            if to_date:
                payments = payments.filter(created_at__date__lte=parse_date(to_date))

        return payments

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Selected filters
        context["selected_user"] = self.request.GET.get("user")
        context["selected_bank"] = self.request.GET.get("bank")
        context["from_date"] = self.request.GET.get("from_date")
        context["to_date"] = self.request.GET.get("to_date")

        # Users dropdown
        if user.is_superuser:
            context["users"] = User.objects.all().order_by("short_name")
            context["banks"] = Bank.objects.all().order_by("name")
            context["total_balance"] = (
                User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            )
        else:
            context["users"] = User.objects.filter(id=user.id)
            context["banks"] = Bank.objects.all().order_by("name")
            context["total_balance"] = user.balance

        return context


@login_required
@user_passes_test(lambda u: u.is_superuser)
def payment_detail(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)
    
    if request.method == "POST":
        form = PaymentApprovalForm(request.POST, instance=payment)
        if form.is_valid():
            payment = form.save(commit=False)
            action = request.POST.get("action")
            user = payment.user
            
            if action == "approve":
                payment.status = "approved"
                payment.approved_by = request.user
                payment.approved_at = now()
                payment.is_office_management = False  # Normal approval
                payment.save()  # This will trigger automatic balance update for user
                
                messages.success(
                    request,
                    f"Payment approved. ${payment.amount} added to {user.email}'s balance.",
                )
                
                subject = "Your Payment Has Been Approved - HFall Realty"
                from_email = settings.DEFAULT_FROM_EMAIL
                to_email = [user.email]
                context = {
                    "user": user,
                    "payment": payment,
                    "date": now().strftime("%d %b %Y"),
                    "status": "approved",
                }
                html_content = render_to_string("emails/user_payment_status.html", context)
                text_content = f"""
                    Dear {user.get_full_name()},
                    Your payment of ${payment.amount} has been approved and added to your balance.
                    Thank you for your contribution.
                    Date: {context['date']}
                    Regards,
                    HFall Finance Team
                """.strip()
                email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
                email.attach_alternative(html_content, "text/html")
                email.send()
                
            elif action == "approve_office":
                # Get the office manager
                office_manager = User.get_office_manager()
                
                if not office_manager:
                    messages.error(request, "No office manager found. Please assign an office manager first.")
                    return redirect("accounts:payment_detail", payment_id=payment.id)
                
                # Manually update payment status WITHOUT triggering save() balance logic
                with transaction.atomic():
                    Payment.objects.filter(pk=payment.pk).update(
                        status='approved',
                        approved_by=request.user,
                        approved_at=now(),
                        is_office_management=True  # Mark as office management
                    )
                    
                    # Only update office manager's balance
                    office_manager.balance += payment.amount
                    office_manager.total_invest_balance += payment.amount
                    office_manager.save()
                
                messages.success(
                    request,
                    f"Payment approved for office management. ${payment.amount} added to office manager ({office_manager.get_full_name()})'s balance. User {user.email}'s balance remains unchanged.",
                )
                
                # Send email to payment submitter
                subject = "Your Payment Has Been Approved (Office Management) - HFall Realty"
                from_email = settings.DEFAULT_FROM_EMAIL
                to_email = [user.email]
                context = {
                    "user": user,
                    "payment": payment,
                    "date": now().strftime("%d %b %Y"),
                    "status": "approved",
                    "office_management": True,
                }
                html_content = render_to_string("emails/user_payment_status.html", context)
                text_content = f"""
                    Dear {user.get_full_name()},
                    Your payment of ${payment.amount} has been approved and added to the office management account.
                    Thank you for your contribution.
                    Date: {context['date']}
                    Regards,
                    HFall Finance Team
                """.strip()
                email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
                email.attach_alternative(html_content, "text/html")
                email.send()
                
            elif action == "reject":
                payment.status = "rejected"
                payment.is_office_management = False
                payment.save()
                
                messages.warning(request, f"Payment rejected for {user.email}.")
                
                subject = "Your Payment Has Been Rejected - HFall Realty"
                from_email = settings.DEFAULT_FROM_EMAIL
                to_email = [user.email]
                context = {
                    "user": user,
                    "payment": payment,
                    "date": now().strftime("%d %b %Y"),
                    "status": "rejected",
                }
                html_content = render_to_string("emails/user_payment_status.html", context)
                text_content = f"""
                    Dear {user.get_full_name()},
                    Unfortunately, your payment of ${payment.amount} has been rejected after review.
                    Please log in to your dashboard and contact support for clarification if needed.
                    Date: {context['date']}
                    Regards,
                    HFall Finance Team
                """.strip()
                email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
                email.attach_alternative(html_content, "text/html")
                email.send()
                
            elif action == "clarification":
                clarification_message = request.POST.get("clarification_message", "").strip()
                
                subject = f"Clarification Needed for Payment #{payment.id}"
                from_email = settings.DEFAULT_FROM_EMAIL
                to_email = [user.email]
                context = {
                    "user_full_name": user.get_full_name() or user.first_name,
                    "payment_id": payment.id,
                    "clarification_message": clarification_message,
                    "date": now().strftime("%d %b %Y"),
                }
                html_content = render_to_string("emails/payment_clarification.html", context)
                text_content = f"""
                    Dear {context['user_full_name']},
                    We need clarification for Payment ID #{payment.id}.
                    Message from admin: {clarification_message}
                    Please log in and update the necessary details or contact support.
                    Regards,
                    HFall Finance Team
                """.strip()
                email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
                email.attach_alternative(html_content, "text/html")
                email.send()
                
                messages.info(request, f"Clarification request email sent to {user.email}.")
                return redirect("accounts:payment_detail", payment_id=payment.id)
            
            return redirect("accounts:pending_payments")
    else:
        form = PaymentApprovalForm(instance=payment)
    
    # Get office manager for template context
    office_manager = User.get_office_manager()
    
    context = {
        "payment": payment, 
        "form": form,
        "office_manager": office_manager,
    }
    
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    
    return render(request, "payment_detail.html", context)



@login_required
def office_management_payments(request):
    """
    View for office manager to see all payments approved for office management
    Only accessible by user with office_management=True
    """
    # Check if current user is office manager
    if not request.user.office_management:
        messages.error(request, "You don't have permission to access this page. Only office manager can view this.")
        return redirect('accounts:dashboard')
    
    # Get all office management payments
    payments = Payment.objects.filter(
        is_office_management=True,
        status='approved'
    ).select_related('user', 'bank', 'approved_by').order_by('-approved_at')
    
    # Calculate total amount
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Get filter parameters
    user_id = request.GET.get('user')
    bank_id = request.GET.get('bank')
    
    # Apply filters
    if user_id:
        payments = payments.filter(user_id=user_id)
    if bank_id:
        payments = payments.filter(bank_id=bank_id)
    
    # Recalculate total after filter
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'payments': payments,
        'total_amount': total_amount,
        'users': User.objects.all(),
        'banks': Bank.objects.all(),
        'office_manager': request.user,
    }
    
    return render(request, 'office_management_payments.html', context)

@login_required
def office_management_dashboard(request):
    """
    Combined view for office manager - shows both payments and deductions
    Only accessible by user with office_management=True
    """
    # Check if current user is office manager
    if not request.user.office_management:
        messages.error(request, "You don't have permission to access this page. Only office manager can view this.")
        return redirect('accounts:dashboard')
    
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    user_id = request.GET.get('user')
    
    # Get Office Management Payments
    payments = Payment.objects.filter(
        is_office_management=True,
        status='approved'
    ).select_related('user', 'bank', 'approved_by')
    
    # Get Deduction History
    deductions = DeductionHistory.objects.filter(
        office_manager=request.user
    ).select_related(
        'property', 
        'user', 
        'property_contribution'
    )
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            payments = payments.filter(approved_at__date__gte=date_from_obj)
            deductions = deductions.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            payments = payments.filter(approved_at__date__lte=date_to_obj)
            deductions = deductions.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    if user_id:
        payments = payments.filter(user_id=user_id)
        deductions = deductions.filter(user_id=user_id)
    
    # Ordering
    payments = payments.order_by('-approved_at')
    deductions = deductions.order_by('-created_at')
    
    # Calculate totals
    total_payments = payments.aggregate(total=Sum('amount'))['total'] or 0
    total_deductions = deductions.aggregate(total=Sum('deduction_amount'))['total'] or 0
    total_income = total_payments + total_deductions
    
    # Get summary by user
    payment_summary = payments.values(
        'user__id', 
        'user__first_name', 
        'user__last_name',
        'user__short_name'
    ).annotate(
        total_amount=Sum('amount'),
        payment_count=Count('id')
    ).order_by('-total_amount')
    
    deduction_summary = deductions.values(
        'user__id', 
        'user__first_name', 
        'user__last_name',
        'user__short_name'
    ).annotate(
        total_deduction=Sum('deduction_amount'),
        deduction_count=Count('id')
    ).order_by('-total_deduction')
    
    context = {
        'payments': payments,
        'deductions': deductions,
        'total_payments': total_payments,
        'total_deductions': total_deductions,
        'total_income': total_income,
        'payment_summary': payment_summary,
        'deduction_summary': deduction_summary,
        'users': User.objects.all(),
        'office_manager': request.user,
        'date_from': date_from,
        'date_to': date_to,
        'selected_user': user_id,
    }
    
    return render(request, 'office_management_dashboard.html', context)
@login_required
def stripe_payment(request):
    if request.method == "POST":
        amount = request.POST.get("amount")
        notes = request.POST.get("notes")
        if not amount:
            messages.error(request, "Please enter a valid amount")
            return redirect("accounts:payment_banks")

        try:
            amount = float(amount)

            fee = round((amount * 0.03) + 0.30, 2)
            total_charge = round(amount + fee, 2) 
            amount_in_cents = int(total_charge * 100)  

            success_url = request.build_absolute_uri(reverse("accounts:payment_success"))
            cancel_url = request.build_absolute_uri(reverse("accounts:payment_banks"))

            checkout_session = stripe.checkout.Session.create(
                payment_method_types=["card"],
                line_items=[
                    {
                        "price_data": {
                            "currency": "usd",
                            "product_data": {"name": f"Account Deposit - ${amount}"},
                            "unit_amount": amount_in_cents, 
                        },
                        "quantity": 1,
                    },
                ],
                mode="payment",
                success_url=(
                    f"{success_url}?session_id={{CHECKOUT_SESSION_ID}}"
                    f"&amount={amount}&fee={fee}&total={total_charge}"
                ),
                cancel_url=cancel_url,
                metadata={
                    "user_id": request.user.id,
                    "notes": notes,
                },
            )

            return redirect(checkout_session.url, code=303)

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect("accounts:payment_banks")

    return render(request, "stripe_payment.html", {"stripe_public_key": settings.STRIPE_PUBLIC_KEY})
@login_required
def payment_success(request):
    session_id = request.GET.get("session_id")
    amount = request.GET.get("amount")  
    fee = request.GET.get("fee")
    total = request.GET.get("total")  

    if not session_id:
        messages.error(request, "Invalid payment session")
        return redirect("accounts:dashboard")

    try:
        session = stripe.checkout.Session.retrieve(session_id)
        notes = session.metadata.get("notes", "")

        if session.payment_status == "paid":
            stripe_bank, _ = Bank.objects.get_or_create(
                name="Stripe",
                defaults={
                    "account_details": "Online payment via Stripe",
                    "is_active": True,
                    "is_card": True,
                },
            )

           
            payment = Payment.objects.create(
                user=request.user,
                bank=stripe_bank,
                amount=float(amount),
                paid_amount=float(total),
                status="pending",
                notes=notes,
            )
            user=payment.user
            subject = f"New Stripe Payment (Pending) - {request.user.get_full_name()}"
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = [user.email]

            context = {
                "user": request.user,
                "bank": stripe_bank,
                "payment": payment,
                "date": now().strftime("%d %b %Y"),
            }

            html_content = render_to_string("emails/admin_payment_notification.html", context)
            text_content = (
                f"A new Stripe payment has been submitted by {request.user.get_full_name()} "
                f"({request.user.email}) and is pending review.\n"
                f"Amount: ${payment.amount}\n"
                f"Bank: {stripe_bank.name}\n"
                f"Date: {context['date']}\n\n"
                "Please review and approve it via the admin panel.\n"
                "HFall Finance System"
            )

            email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
            email.attach_alternative(html_content, "text/html")
            email.send()

           
            messages.success(
                request,
                f"Payment successful! You sent ${total}. Fee: ${fee}. "
                f"${amount} will be added to your balance after admin review."
            )

            return render(
                request,
                "payment_success.html",
                {
                    "original_amount": amount,
                    "card_fee": fee,
                    "net_amount": total,
                }
            )
        else:
            messages.error(request, "Payment was not completed")
            return redirect("accounts:payment_banks")

    except Exception as e:
        messages.error(request, f"Error verifying payment: {str(e)}")
        return redirect("accounts:dashboard")



def stripe_webhook(request):
    webhook_secret = settings.STRIPE_WEBHOOK_SECRET
    payload = request.body
    sig_header = request.META.get("HTTP_STRIPE_SIGNATURE")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, webhook_secret)
    except ValueError as e:
        # Invalid payload
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        # Invalid signature
        return HttpResponse(status=400)

    # Handle the event
    if event["type"] == "checkout.session.completed":
        session = event["data"]["object"]

        # Retrieve the user from metadata
        user_id = session.get("metadata", {}).get("user_id")
        if user_id:

            pass

    return HttpResponse(status=200)

def get_square_client():
   
    return Client(
        access_token=settings.SQUARE_ACCESS_TOKEN,
        environment=settings.SQUARE_ENVIRONMENT
    )


@login_required
def square_create_payment(request):
    if request.method == "POST":
        amount = request.POST.get("amount")
        notes = request.POST.get("notes", "")
        if not amount:
            messages.error(request, "Please enter a valid amount")
            return redirect("accounts:payment_banks")

        try:
            amount = float(amount)
            fee = round((amount * 0.03) + 0.30, 2)
            total_charge = round(amount + fee, 2)
            amount_in_cents = int(total_charge * 100) 

            client = Client(
                access_token=settings.SQUARE_ACCESS_TOKEN,
                environment=settings.SQUARE_ENVIRONMENT
            )
            checkout_api = client.checkout

            body = {
                "idempotency_key": str(uuid.uuid4()),
                "order": {
                    "order": {
                        "location_id": settings.SQUARE_LOCATION_ID,
                        "line_items": [
                            {
                                "name": f"Account Deposit: ${amount}",
                                "quantity": "1",
                                "base_price_money": {
                                    "amount": amount_in_cents,  
                                    "currency": "USD"
                                }
                            }
                        ]
                    }
                },
                "ask_for_shipping_address": False,
                "pre_populate_buyer_email": request.user.email,
                "redirect_url": request.build_absolute_uri(reverse('accounts:square_payment_success'))
            }

            response = checkout_api.create_checkout(
                location_id=settings.SQUARE_LOCATION_ID, 
                body=body
            )

            if response.is_success():
                checkout_url = response.body['checkout']['checkout_page_url']
                request.session['square_payment_data'] = {
                    'original_amount': str(amount),
                    'fee': str(fee),
                    'total_charge': str(total_charge),
                    'notes': notes,
                }
                return redirect(checkout_url)
            else:
                messages.error(request, "Error creating Square checkout")
                return redirect("accounts:payment_banks")

        except Exception as e:
            messages.error(request, f"Error creating Square payment: {str(e)}")
            return redirect("accounts:payment_banks")

    return redirect("accounts:payment_banks")

@login_required
def square_payment_success(request):
    payment_data = request.session.get('square_payment_data')
    if not payment_data:
        messages.error(request, "Payment session expired")
        return redirect("accounts:payment_banks")

    try:
        original_amount = float(payment_data.get("original_amount", "0"))
        fee = float(payment_data.get("fee", "0"))
        total_charge = float(payment_data.get("total_charge", "0"))
        notes = payment_data.get("notes", "")

        square_bank, _ = Bank.objects.get_or_create(
            name="Square",
            defaults={
                "account_details": "Online payment via Square",
                "is_active": True,
                "is_square": True,
            },
        )

        payment = Payment.objects.create(
            user=request.user,
            bank=square_bank,
            amount=original_amount, 
            paid_amount=total_charge,
            status="pending",
            notes=notes,
        )

        # Email notification (same as Stripe)
        user=payment.user
        subject = f"New Square Payment (Pending) - {request.user.get_full_name()}"
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [user.email]

        context = {
            "user": request.user,
            "payment": payment,
            "date": now().strftime("%d %b %Y"),
        }

        html_content = render_to_string("emails/admin_payment_notification.html", context)
        text_content = (
            f"A new Square payment has been submitted by {request.user.get_full_name()} "
            f"({request.user.email}) and is pending review.\n"
            f"Amount: ${payment.amount}\n"
            f"Date: {context['date']}\n\n"
            "Please review and approve it via the admin panel.\n"
            "HFall Finance System"
        )

        email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
        email.attach_alternative(html_content, "text/html")
        email.send()

        messages.success(
            request,
            f"Square payment successful! You sent ${total_charge:.2f}. Fee: ${fee:.2f}. "
            f"${original_amount:.2f} will be added to your balance after admin review."
        )

        del request.session['square_payment_data']

        return render(request, "payment_success.html", {
            "original_amount": f"{original_amount:.2f}",
            "card_fee": f"{fee:.2f}",
            "net_amount": f"{total_charge:.2f}",
            "payment_method": "Square"
        })

    except Exception as e:
        messages.error(request, f"Error processing Square payment: {str(e)}")
        return redirect("accounts:payment_banks")


paypalrestsdk.configure({
    "mode": "live",  # sandbox or live
    "client_id": "AaUpRxa9MVVgAz-4-Mr262QMqLy9pb-MOKIY_82FM7Hlg56GpwjW6PA9e7Ow784B38HneBAFxn9i1HoC",
    "client_secret": "EP2IKcRnf6GfwkdIKCpeCNzkC_hq3dc_HnReJ8fIaHREnHG_xjwkghm6RQc1UCT85tfJ4DQL9L99Wg8Q"
})


@login_required
def paypal_create_payment(request):
    if request.method == "POST":
        amount = request.POST.get("amount")
        notes = request.POST.get("notes", "")
        if not amount:
            messages.error(request, "Please enter a valid amount")
            return redirect("accounts:payment_banks")

        try:
            # PayPal payment create
            payment = paypalrestsdk.Payment({
                "intent": "sale",
                "payer": {"payment_method": "paypal"},
                "redirect_urls": {
                    "return_url": request.build_absolute_uri(reverse("accounts:paypal_success")),
                    "cancel_url": request.build_absolute_uri(reverse("accounts:payment_banks")),
                },
                "transactions": [{
                    "item_list": {
                        "items": [{
                            "name": f"Account Deposit - ${amount}",
                            "sku": "deposit",
                            "price": str(amount),
                            "currency": "USD",
                            "quantity": 1,
                        }]
                    },
                    "amount": {"total": str(amount), "currency": "USD"},
                    "description": notes,
                }]
            })

            if payment.create():
                for link in payment.links:
                    if link.rel == "approval_url":
                        approval_url = str(link.href)
                        request.session['paypal_payment_data'] = {
                            "payment_id": payment.id,
                            "amount": amount,
                            "notes": notes,
                        }
                        return redirect(approval_url)
            else:
                messages.error(request, "Error creating PayPal payment")
                return redirect("accounts:payment_banks")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect("accounts:payment_banks")

    return redirect("accounts:payment_banks")


@login_required
def paypal_success(request):
    payment_data = request.session.get("paypal_payment_data")
    if not payment_data:
        messages.error(request, "Payment session expired")
        return redirect("accounts:payment_banks")

    payment_id = request.GET.get("paymentId")
    payer_id = request.GET.get("PayerID")

    if not payment_id or not payer_id:
        messages.error(request, "Invalid PayPal response")
        return redirect("accounts:payment_banks")

    try:
        payment = paypalrestsdk.Payment.find(payment_id)
        if payment.execute({"payer_id": payer_id}):
            paypal_bank, _ = Bank.objects.get_or_create(
                name="PayPal",
                defaults={
                    "account_details": "Online payment via PayPal",
                    "is_active": True,
                    "is_paypal": True,
                },
            )
            new_payment = Payment.objects.create(
                user=request.user,
                bank=paypal_bank,
                amount=float(payment_data["amount"]),
                paid_amount=float(payment_data["amount"]), 
                status="pending",
                notes=payment_data.get("notes", ""),
            )

            # Email to admin
            user=payment.user
            subject = f"New PayPal Payment (Pending) - {request.user.get_full_name()}"
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = [user.email]
            context = {
                "user": request.user,
                "payment": new_payment,
                "date": now().strftime("%d %b %Y"),
            }
            html_content = render_to_string("emails/admin_payment_notification.html", context)
            text_content = f"A new PayPal payment has been submitted by {request.user.get_full_name()}"

            email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
            email.attach_alternative(html_content, "text/html")
            email.send()

            messages.success(
                request,
                f"PayPal payment successful! You sent ${payment_data['amount']}. "
                f"${payment_data['amount']} will be added to your balance after admin review."
            )
            del request.session['paypal_payment_data']
            return render(request, "payment_success.html", {
                "original_amount": payment_data["amount"],
                "payment_method": "PayPal",
            })
        else:
            messages.error(request, "Payment was not completed")
            return redirect("accounts:payment_banks")
    except Exception as e:
        messages.error(request, f"Error processing PayPal payment: {str(e)}")
        return redirect("accounts:payment_banks")


def get_square_client():
    return Client(
        access_token=settings.SQUARE_ACCESS_TOKEN,
        environment=settings.SQUARE_ENVIRONMENT,
    )

@login_required
def square_checkout(request):
    amount = request.GET.get("amount", "")
    notes = request.GET.get("notes", "")

    context = {
        "square_app_id": settings.SQUARE_APPLICATION_ID,
        "square_location_id": settings.SQUARE_LOCATION_ID,
        "environment": settings.SQUARE_ENVIRONMENT,
        "prefill_amount": amount,
        "prefill_notes": notes,
    
        "success_url": reverse("accounts:square_success"),
    }
    return render(request, "square_checkout.html", context)

@method_decorator(login_required, name="dispatch")
class expense_list(ListView):
    model = Expense
    template_name = "expenses/expense_list.html"
    context_object_name = "expenses"
    paginate_by = 100000

    def get_queryset(self):
        expenses = Expense.objects.filter(
            property__isnull=False
        ).select_related('property', 'created_by', 'approved_by').order_by("-expense_date")

        # Get filters
        status_filter = self.request.GET.get("status", "")
        created_by_filter = self.request.GET.get("created_by", "")
        from_date = self.request.GET.get("from_date", "")
        to_date = self.request.GET.get("to_date", "")

        # Apply filters
        if status_filter:
            expenses = expenses.filter(status=status_filter)

        if created_by_filter:
            expenses = expenses.filter(created_by_id=created_by_filter)

        if from_date:
            expenses = expenses.filter(expense_date__gte=parse_date(from_date))

        if to_date:
            expenses = expenses.filter(expense_date__lte=parse_date(to_date))

        return expenses

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Grab filters again for context
        status_filter = self.request.GET.get("status", "")
        created_by_filter = self.request.GET.get("created_by", "")
        from_date = self.request.GET.get("from_date", "")
        to_date = self.request.GET.get("to_date", "")

        context.update({
            "status_filter": status_filter,
            "created_by_filter": created_by_filter,
            "from_date": from_date,
            "to_date": to_date,
            "users": User.objects.all().order_by("short_name"),  # For dropdown
        })
        
        filtered_expenses = self.get_queryset()
        total_expense_amount = filtered_expenses.aggregate(
            total=Sum("amount")
        )["total"] or 0

        context["total_expense_amount"] = total_expense_amount

        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance

        return context

@csrf_exempt
@require_POST
@login_required
def update_expense_status(request):
    expense_id = request.POST.get("expense_id")
    new_status = request.POST.get("new_status")

    expense = get_object_or_404(Expense, id=expense_id)

    success = expense.update_status(new_status, request.user)

    return JsonResponse({
        "success": success,
        "new_status": expense.get_status_display(),
        "approved_by": expense.approved_by.get_full_name() if expense.approved_by else "-",
    })
@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == "POST":
        expense.delete()
        messages.success(request, "Expense deleted successfully ✅")
        return redirect("accounts:expense_list")
    return redirect("accounts:expense_list") 
@login_required
def expense_download(request):
    """Download expenses as Excel file using openpyxl"""
    # Get status filter parameter from request
    status_filter = request.GET.get("status", "")

    # Base query for all expenses
    expenses = Expense.objects.all().order_by("-created_at")

    # Apply filter if status is provided
    if status_filter:
        expenses = expenses.filter(status=status_filter)
        filename_status = f"_{status_filter}"
    else:
        filename_status = ""

    # Create a new workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Expenses"

    # Define header style
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )

    # Add header row
    headers = [
        "Purpose",
        "Description",
        "Amount",
        "Status",
        "Created By",
        "Expense Date",
        "Created At",
        "Property",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add data rows
    for row_idx, expense in enumerate(expenses, start=2):
        # Column 1: Purpose
        worksheet.cell(row=row_idx, column=1, value=expense.purpose or "")

        # Column 2: Description
        worksheet.cell(row=row_idx, column=2, value=expense.description or "")

        # Column 3: Amount
        worksheet.cell(
            row=row_idx, column=3, value=float(expense.amount) if expense.amount else 0
        )

        # Column 4: Status
        worksheet.cell(row=row_idx, column=4, value=expense.get_status_display())

        # Column 5: Created By
        worksheet.cell(
            row=row_idx,
            column=5,
            value=expense.created_by.get_full_name() if expense.created_by else "",
        )

        # Column 6: Expense Date
        worksheet.cell(
            row=row_idx,
            column=6,
            value=(
                expense.expense_date.strftime("%Y-%m-%d")
                if expense.expense_date
                else ""
            ),
        )

        # Column 7: Created At
        worksheet.cell(
            row=row_idx,
            column=7,
            value=(
                expense.created_at.strftime("%Y-%m-%d %H:%M")
                if expense.created_at
                else ""
            ),
        )

        # Column 8: Property
        worksheet.cell(
            row=row_idx,
            column=8,
            value=expense.property.property_name if expense.property else "",
        )

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save workbook to BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Set up the response
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"{current_date}{filename_status}_expenses.xlsx"

    # Create HttpResponse with appropriate headers
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response

@login_required
def expense_create(request):
    """Create a new expense request"""
    if not request.user.is_superuser and not request.user.is_finnancial and not request.user.is_expense:
        messages.error(request, "Only financial users or admins can create expenses.")
        return redirect("accounts:expense_list")
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.created_by = request.user
            expense.save()
            messages.success(request, "Expense request submitted successfully.")
            return redirect("accounts:expense_list")
    else:
        form = ExpenseForm()
    context = {"form": form}
    # Total balance logic
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "expenses/expense_form.html", context)


@login_required
def expense_update(request, expense_id):
    """Update an existing expense request"""
    expense = get_object_or_404(Expense, id=expense_id)

    # Permission check
    if (request.user != expense.created_by and not request.user.is_superuser and not request.user.is_expense):
        messages.error(request, "You do not have permission to edit this expense.")
        return redirect("accounts:expense_list")

    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense updated successfully.")
            return redirect("accounts:expense_list")
    else:
        form = ExpenseForm(instance=expense)

    context = {"form": form}

    # Total balance logic
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance

    return render(request, "expenses/expense_form.html", context)


@login_required
def expense_detail(request, expense_id):
    """View expense details"""
    expense = get_object_or_404(Expense, id=expense_id)

    context = {"expense": expense}
    # Total balance logic
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "expenses/expense_detail.html", context)


@login_required
def expense_approve(request, expense_id):
    """Approve an expense request - only superusers can do this"""
    expense = get_object_or_404(Expense, id=expense_id)

    if expense.status != "pending":
        messages.error(request, "This expense has already been processed.")
        return redirect("accounts:expense_detail", expense_id=expense.id)

    # Permission check
    if not request.user.is_superuser and not request.user.is_finnancial:
        messages.error(request, "You do not have permission to reject this expense.")
        return redirect("accounts:expense_list")

    # Creator cannot reject own expense
    if expense.created_by == request.user:
        messages.error(request, "You cannot reject your own expense.")
        return redirect("accounts:expense_list")

    if expense.approve_expense(request.user):
        try:
            # ✅ Send approval email to user
            user = expense.created_by
            subject = f"Expense #{expense.id} Approved - HFall Realty"
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = [user.email]

            context = {
                "user_full_name": user.get_full_name() or user.first_name,
                "expense_id": expense.id,
                "date": now().strftime("%d %b %Y"),
                "status": "approved",
            }

            html_content = render_to_string("emails/expense_status_update.html", context)
            text_content = f"""
                Dear {context['user_full_name']},

                Your expense (ID: {expense.id}) has been approved.

                Date: {context['date']}

                Regards,
                HFall Finance Team
                """.strip()

            email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
            email.attach_alternative(html_content, "text/html")
            email.send()

            messages.success(request, "Expense approved successfully and email sent.")

        except Exception as e:
            messages.warning(request, f"Expense approved but email could not be sent: {e}")
    else:
        messages.error(request, "Error approving expense.")

    return redirect("accounts:expense_list")


@login_required
def expense_reject(request, expense_id):
    """Reject an expense request - only superusers can do this"""
    expense = get_object_or_404(Expense, id=expense_id)

    if expense.status != "pending":
        messages.error(request, "This expense has already been processed.")
        return redirect("accounts:expense_detail", expense_id=expense.id)

    # Permission check
    if not request.user.is_superuser and not request.user.is_finnancial:
        messages.error(request, "You do not have permission to reject this expense.")
        return redirect("accounts:expense_list")

    # Creator cannot reject own expense
    if expense.created_by == request.user:
        messages.error(request, "You cannot reject your own expense.")
        return redirect("accounts:expense_list")

    if expense.reject_expense(request.user):
      
        user = expense.created_by
        subject = f"Expense #{expense.id} Rejected - HFall Realty"
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [user.email]

        context = {
            "user_full_name": user.get_full_name() or user.first_name,
            "expense_id": expense.id,
            "date": now().strftime("%d %b %Y"),
            "status": "rejected",
        }

        html_content = render_to_string("emails/expense_status_update.html", context)
        text_content = f"""
            Dear {context['user_full_name']},

            Your expense (ID: {expense.id}) has been rejected.

            Date: {context['date']}

            Please log in for more details or contact the finance team if needed.

            Regards,
            HFall Finance Team
            """.strip()

        email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
        email.attach_alternative(html_content, "text/html")
        email.send()

        messages.success(request, "Expense rejected successfully.")
    else:
        messages.error(request, "Error rejecting expense.")

    return redirect("accounts:expense_list")



@login_required
def expense_clarification(request, expense_id):
    """
    Send a clarification email to the user who created the expense.
    """
    expense = get_object_or_404(Expense, id=expense_id)
    user = expense.created_by

    if request.method == "POST":
        subject = f"Clarification Needed for Expense #{expense.id}"
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = [user.email]

        # get message from form
        clarification_message = request.POST.get("clarification_message", "").strip()

        context = {
            "user_full_name": user.get_full_name() or user.first_name,
            "expense_id": expense.id,
            "date": now().strftime("%d %b %Y"),
            "clarification_message": clarification_message,
        }

        html_content = render_to_string("emails/expense_clarification.html", context)
        text_content = (
            f"Dear {context['user_full_name']},\n\n"
            f"We need clarification for Expense ID {expense.id}.\n\n"
            f"Message:\n{clarification_message}\n\n"
            "Please log in and update it.\n\nRegards,\nHFall Finance Team"
        )

        email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
        email.attach_alternative(html_content, "text/html")


        email.send()

        messages.success(request, f"Clarification email sent to {user.email}.")
        return redirect("accounts:expense_detail", expense_id=expense.id)

    return redirect("accounts:expense_detail", expense_id=expense.id)

# views.py - Updated payment view and list view

@login_required
def expense_payment_create(request, expense_id):
    if not request.user.is_superuser and not getattr(request.user, "is_finnancial", False):
        messages.error(request, "Only financial users or admins can pay expenses.")
        return redirect("accounts:expense_payment_list")
    
    expense = get_object_or_404(Expense, id=expense_id, status="approved")
    
    # Check if this expense should be in the payment list
    if not expense.should_add_to_expense_balance():
        messages.error(request, "This expense is not eligible for payment (property expense without user payment).")
        return redirect("accounts:expense_payment_list")
    
    # Pre-fill user field if paid_by_user is set
    initial_data = {
        "amount": Decimal(expense.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    }
    if expense.paid_by_user:
        initial_data["user"] = expense.paid_by_user
    
    form = ExpensePaymentForm(
        request.POST or None,
        request.FILES or None,
        initial=initial_data
    )
    form.instance.expense = expense  
    
    if request.method == "POST":
        print("POST request received")  
        if form.is_valid():
            payment = form.save(commit=False)
            payment.amount = Decimal(expense.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            payment.process_payment() 
            print("Payment processed")  
            messages.success(request, "Expense payment processed successfully.")
            return redirect("accounts:expense_payment_list")
        else:
            print("Form errors:", form.errors)  
    
    return render(request, "expenses/expense_payment_form.html", {"form": form, "expense": expense})


# @method_decorator(login_required, name="dispatch")
# class expense_payment_list(ListView):
#     model = Expense
#     template_name = "expenses/user_expense_paylist.html"
#     context_object_name = "expenses"
#     paginate_by = 1000000

#     def get_queryset(self):
#         """
#         Only show approved expenses that should be paid:
#         1. Non-property expenses (always need payment), OR
#         2. Property expenses with paid_by_user set (need reimbursement)
#         """
#         from django.db.models import Q
        
#         # Get all approved expenses
#         queryset = Expense.objects.filter(status="approved").select_related(
#             'paid_by_user', 'property', 'created_by', 'approved_by'
#         ).order_by("-created_at")
        
#         payable_expenses = queryset.filter(
#             Q(property__isnull=True) |  # Non-property expenses
#             Q(property__isnull=False, paid_by_user__isnull=False)  # Property expenses with user
#         )
        
#         return payable_expenses

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)

#         # Total balance logic
#         if self.request.user.is_superuser:
#             total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
#             context["total_balance"] = total_investment
#         else:
#             context["total_balance"] = self.request.user.balance
        
#         # Add expense balance for display
#         from .models import ExpenseBalance
#         context["expense_balance"] = ExpenseBalance.objects.filter(id=1).first()

#         return context
from django.views.generic import ListView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.db.models import Q, Sum

@method_decorator(login_required, name="dispatch")
class expense_payment_list(ListView):
    model = Expense
    template_name = "expenses/user_expense_paylist.html"
    context_object_name = "expenses"
    paginate_by = 10000

    def get_queryset(self):
        """
        Payable approved expenses:
        1. Non-property expenses
        2. Property expenses with paid_by_user (reimbursement)
        """

        qs = Expense.objects.filter(
            status="approved"
        ).select_related(
            "property", "paid_by_user", "created_by"
        ).order_by("-expense_date")

        qs = qs.filter(
            Q(property__isnull=True) |
            Q(property__isnull=False, paid_by_user__isnull=False)
        )

        # -------- Filters --------
        payment_status = self.request.GET.get("payment_status", "")
        paid_by = self.request.GET.get("paid_by", "")
        property_id = self.request.GET.get("property", "")

        if payment_status:
            qs = qs.filter(payment_status=payment_status)

        if paid_by:
            qs = qs.filter(paid_by_user_id=paid_by)

        if property_id:
            qs = qs.filter(property_id=property_id)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        filtered_qs = self.get_queryset()

        # Payable SUM
        context["total_payable"] = filtered_qs.aggregate(
            total=Sum("amount")
        )["total"] or 0

        # Filters data
        context.update({
            "users": User.objects.all().order_by("short_name"),
            "properties": Property.objects.all().order_by("title"),
            "payment_status_filter": self.request.GET.get("payment_status", ""),
            "paid_by_filter": self.request.GET.get("paid_by", ""),
            "property_filter": self.request.GET.get("property", ""),
        })

        # Balance info
        if self.request.user.is_superuser:
            context["total_balance"] = User.objects.aggregate(
                Sum("balance")
            )["balance__sum"] or 0
        else:
            context["total_balance"] = self.request.user.balance

        return context



@method_decorator(login_required, name="dispatch")
class ExpenseDetailView(DetailView):
    model = Expense
    template_name = "expenses/expense_details.html"
    context_object_name = "expense"
@login_required
def expense_payment_create(request, expense_id):
    if not request.user.is_superuser and not getattr(request.user, "is_finnancial", False):
        messages.error(request, "Only financial users or admins can pay expenses.")
        return redirect("accounts:expense_payment_list")
    
    expense = get_object_or_404(Expense, id=expense_id, status="approved")
    
    # Check if this expense should be in the payment list
    if not expense.should_add_to_expense_balance():
        messages.error(request, "This expense is not eligible for payment (property expense without user payment).")
        return redirect("accounts:expense_payment_list")
    
    # Pre-fill user field if paid_by_user is set
    initial_data = {
        "amount": Decimal(expense.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    }
    if expense.paid_by_user:
        initial_data["user"] = expense.paid_by_user
    
    form = ExpensePaymentForm(
        request.POST or None,
        request.FILES or None,
        initial=initial_data
    )
    form.instance.expense = expense  
    
    if request.method == "POST":
        print("POST request received")  
        if form.is_valid():
            payment = form.save(commit=False)
            payment.amount = Decimal(expense.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            payment.process_payment() 
            print("Payment processed")  
            messages.success(request, "Expense payment processed successfully.")
            return redirect("accounts:expense_payment_list")
        else:
            print("Form errors:", form.errors)  
    
    return render(request, "expenses/expense_payment_form.html", {"form": form, "expense": expense})





@login_required
def get_expense_details(request, expense_id):
    """Return expense details for AJAX"""
    expense = get_object_or_404(Expense, id=expense_id, status="approved")
    data = {
        "amount": float(expense.amount),
        "user_id": expense.created_by.id,
        "user_name": expense.created_by.get_full_name() or expense.created_by.username,
    }
    return JsonResponse(data)

def get_client_ip(request):
    """
    Returns the client's IP address from the given request object.
    """
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0].strip()
    else:
        ip = request.META.get("REMOTE_ADDR", None)
    return ip


from django.core.exceptions import PermissionDenied


def upload_new_agreement(request):
    if not request.user.is_authenticated:
        raise PermissionDenied("You must be logged in to upload an agreement.")

    if request.method == "POST" and request.FILES.get("agreement_file"):
        uploaded_file = request.FILES["agreement_file"]

        UserAgreement.objects.create(user=request.user, uploaded_file=uploaded_file)
        messages.success(request, "Agreement uploaded successfully.")
        return redirect("accounts:user_agreement_list")  # or wherever you want
    context = {}
    # Total balance logic
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "upload_agreement.html", context)


class UploadedAgreementsView(LoginRequiredMixin, ListView):
    model = UserAgreement
    template_name = "dashboard/agreement/user_agreement_list.html"
    context_object_name = "uploaded_agreements"
    paginate_by = 20

    def get_queryset(self):
        return (
            UserAgreement.objects.filter(uploaded_file__isnull=False)
            .exclude(uploaded_file="")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance

        return context


class UploadedAgreementsDeleteView(View):
    def delete(self, request, pk, *args, **kwargs):
        agreement = get_object_or_404(UserAgreement, pk=pk)

        # Allow only the owner or a superuser
        if request.user != agreement.user and not request.user.is_superuser:
            return HttpResponseForbidden(
                "You do not have permission to delete this agreement."
            )

        agreement.delete()
        return JsonResponse({"success": True})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = (
                User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            )
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance
        return context


@user_passes_test(lambda u: u.is_superuser)
def create_announcement(request):
    if request.method == "POST":
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.created_by = request.user
            announcement.save()

            users = User.objects.filter(is_superuser=False, is_active=True)
            recipient_list = [user.email for user in users if user.email]

            # Render email content
            context = {"announcement": announcement}
            html_content = render_to_string("emails/announcement_email.html", context)
            text_content = strip_tags(html_content)

            subject = announcement.title
            from_email = settings.DEFAULT_FROM_EMAIL

            # Use EmailMultiAlternatives for HTML email
            email = EmailMultiAlternatives(
                subject, text_content, from_email, recipient_list
            )
            email.attach_alternative(html_content, "text/html")
            email.send()

            return redirect("accounts:dashboard")
    else:
        form = AnnouncementForm()

    context = {"form": form}
    # Total balance logic
    if request.user.is_superuser:
        total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
        context["total_balance"] = total_investment
    else:
        context["total_balance"] = request.user.balance
    return render(request, "create_announcement.html", context)


class ExpensesCopyView(LoginRequiredMixin, View):
    def post(self, request, pk, *args, **kwargs):
        original = get_object_or_404(Expense, pk=pk)
        copied = Expense.objects.create(
            purpose=f"{original.purpose} (Copy)",
            description=original.description,
            # expense_date=original.expense_date,
            expense_date=timezone.now().date(),
            status='pending',
            image=original.image,
            property=original.property, 
            amount=original.amount,
            created_by=request.user,
        )
        return JsonResponse({'status': 'success', 'copied_id': copied.pk})

@user_passes_test(lambda u: u.is_superuser)
def help_create_view(request):
    if request.method == 'POST':
        form = HelpForm(request.POST, request.FILES)
        if form.is_valid():
            help_instance = form.save()

            subject = f"New Help Entry: {help_instance.title}"
            from_email = settings.DEFAULT_FROM_EMAIL
            to_emails = ['malaminmiu@gmail.com', 'tofael483@gmail.com']
            context = {
                'title': help_instance.title,
                'description': help_instance.description,
                'created_at': help_instance.created_at.strftime('%B %d, %Y'),
                'image_cid': 'helpimage'  
            }

            html_content = render_to_string('emails/help_created.html', context)
            text_content = strip_tags(html_content)

            msg = EmailMultiAlternatives(subject, text_content, from_email, to_emails)
            msg.attach_alternative(html_content, "text/html")

            # Attach image if uploaded
            if help_instance.image:
                image_path = help_instance.image.path
                with open(image_path, 'rb') as f:
                    image = MIMEImage(f.read())
                    image.add_header('Content-ID', '<helpimage>')
                    image.add_header('Content-Disposition', 'inline', filename=os.path.basename(image_path))
                    msg.attach(image)

            msg.send()
            return redirect('accounts:help_success')
    else:
        form = HelpForm()
    return render(request, 'help_create.html', {'form': form})

def help_success_view(request):
    return render(request, 'snippet/success.html')



@method_decorator(login_required, name='dispatch')
class PaymentListView(ListView):
    model = Payment
    template_name = 'payment_list.html'
    context_object_name = 'payments'
    paginate_by = 20

    def get_queryset(self):
        queryset = Payment.objects.select_related('user', 'bank', 'approved_by').all()

        status = self.request.GET.get('status', 'approved') 
        user_id = self.request.GET.get('user')
        bank_id = self.request.GET.get('bank')

        if status:
            queryset = queryset.filter(status=status)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if bank_id:
            queryset = queryset.filter(bank_id=bank_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)      
        queryset = self.get_queryset()
        context['statuses'] = Payment.STATUS_CHOICES
        context['users'] = User.objects.all().order_by("short_name")
        context['banks'] = Bank.objects.all().order_by("name")
        context['total_amount'] = queryset.aggregate(total=Sum('amount'))['total'] or 0
        return context



@csrf_exempt
@require_POST
@login_required
def update_payment_status(request):
    payment_id = request.POST.get("payment_id")
    new_status = request.POST.get("new_status")

    payment = get_object_or_404(Payment, id=payment_id)
    payment.status = new_status
    payment.approved_by = request.user if new_status == "approved" else None
    payment.save()

    return JsonResponse({
        "success": True,
        "new_status": payment.get_status_display(),
        "approved_by": payment.approved_by.get_full_name() if payment.approved_by else "-",
        "approved_at": payment.approved_at.strftime("%Y-%m-%d %H:%M") if payment.approved_at else "-"
    })



@method_decorator(login_required, name='dispatch')
class PaymentExportExcelView(View):
    def get(self, request, *args, **kwargs):
        queryset = Payment.objects.select_related('user', 'bank', 'approved_by').all()

        status = request.GET.get('status')
        user_id = request.GET.get('user')
        bank_id = request.GET.get('bank')

        if status:
            queryset = queryset.filter(status=status)
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if bank_id:
            queryset = queryset.filter(bank_id=bank_id)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"

        # Headers
        headers = [
            "User", "Bank", "Amount", "Status",
             "Created At", "Approved By", "Approved At"
        ]
        ws.append(headers)

        # Data rows
        for payment in queryset:
            ws.append([
                str(payment.user),
                str(payment.bank),
                float(payment.amount),
                payment.get_status_display(),
                payment.created_at.strftime("%Y-%m-%d %H:%M"),
                str(payment.approved_by) if payment.approved_by else '',
                payment.approved_at.strftime("%Y-%m-%d %H:%M") if payment.approved_at else ''
            ])

        # Return response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
        wb.save(response)
        return response
    
    

@login_required
def add_beneficiary(request):
    BeneficiaryFormSet = modelformset_factory(Beneficiary, form=BeneficiaryForm, extra=1, can_delete=True)

    if request.method == "POST":
        formset = BeneficiaryFormSet(request.POST, queryset=Beneficiary.objects.none())
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data:
                    beneficiary = form.save(commit=False)
                    beneficiary.user = request.user
                    beneficiary.save()
            messages.success(request, "Beneficiaries added successfully!")
            return redirect('accounts:profile') 
    else:
        formset = BeneficiaryFormSet(queryset=Beneficiary.objects.none())

    return render(request, 'add_beneficiary.html', {'formset': formset})


class OfficeExpenseListView(ListView):
    """
    List of office management expenses that are not property-related.
    Only CREATED expenses are shown (approval is not required).
    """
    model = Expense
    template_name = "expenses/office_expense_list.html"
    context_object_name = "expenses"
    paginate_by = 50

    def dispatch(self, request, *args, **kwargs):
        # Only office manager or superuser can access
        if not (request.user.is_superuser or request.user.office_management):
            messages.error(request, "Only Office Manager can access this page.")
            return redirect("accounts:dashboard")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        """
        All expenses that are not property-related (any status).
        CRITICAL: Approval is not required, only creation matters.
        """
        queryset = Expense.objects.filter(
            property__isnull=True,
        ).select_related('created_by', 'approved_by').order_by("-created_at")
        
        print(f"\n{'='*60}")
        print(f"🔍 OFFICE EXPENSE QUERYSET DEBUG")
        print(f"{'='*60}")
        print(f"Total Non-Property Expenses: {queryset.count()}")
        
        for expense in queryset[:5]:
            print(f"\nExpense ID: {expense.id}")
            print(f"Purpose: {expense.purpose}")
            print(f"Amount: ${expense.amount}")
            print(f"Status: {expense.status}")
            print(f"Payment Status: {expense.payment_status}")
            print(f"Property: {expense.property}")
            print(f"Created At: {expense.created_at}")
        
        print(f"{'='*60}\n")
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Current balance of office manager
        office_manager = User.get_office_manager()
        context["office_manager"] = office_manager
        context["office_manager_balance"] = office_manager.balance if office_manager else Decimal('0')
        
        # Total unpaid expenses (non-property)
        unpaid_total = Expense.objects.filter(
            property__isnull=True,
            payment_status="unpaid"
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal('0')
        context["unpaid_total"] = unpaid_total
        
        # Total paid expenses (non-property)
        paid_total = Expense.objects.filter(
            property__isnull=True,
            payment_status="paid"
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal('0')
        context["paid_total"] = paid_total
        
        # Total number of expenses
        context["total_expenses"] = self.get_queryset().count()
        
        print(f"\n📊 CONTEXT DATA:")
        print(f"Office Manager: {office_manager.get_full_name() if office_manager else 'None'}")
        print(f"Balance: ${context['office_manager_balance']}")
        print(f"Unpaid Total: ${unpaid_total}")
        print(f"Paid Total: ${paid_total}")
        print(f"Total Expenses: {context['total_expenses']}\n")
        
        return context
@login_required
def office_expense_payment(request, expense_id):
    """
    Office expense payment processing.
    User can be selected to make the payment.
    Deducts from the office manager's balance.
    """
    # Permission check
    if not (request.user.is_superuser or request.user.office_management):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect("accounts:office_expense_list")
    
    # Fetch the expense - no status check
    expense = get_object_or_404(
        Expense, 
        id=expense_id, 
        property__isnull=True,  # Only non-property expenses
        payment_status="unpaid"  # Only unpaid expenses
    )
    
    # Fetch the office manager
    office_manager = User.get_office_manager()
    if not office_manager:
        messages.error(request, "No Office Manager found. Please assign one first.")
        return redirect("accounts:office_expense_list")
    
    # Initialize form
    initial_data = {
        "amount": Decimal(expense.amount).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    }

  
    if expense.paid_by_user:
        initial_data["user"] = expense.paid_by_user

    form = OfficeExpensePaymentForm(
        request.POST or None,
        request.FILES or None,
        initial=initial_data
    )

    form.instance.expense = expense
    
    if request.method == "POST":
        print("\n" + "="*60)
        print("📝 POST REQUEST RECEIVED")
        print("="*60)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    selected_user = form.cleaned_data['user']
                    office_manager = User.objects.select_for_update().get(id=office_manager.id)
                    expense.refresh_from_db()
                    expense_amount = Decimal(expense.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    
                    print(f"\n💰 OFFICE EXPENSE PAYMENT PROCESSING")
                    print(f"{'='*60}")
                    print(f"Expense ID: {expense.id}")
                    print(f"Purpose: {expense.purpose}")
                    print(f"Amount: ${expense_amount}")
                    print(f"Status: {expense.status}")
                    print(f"Payment Status: {expense.payment_status}")
                    print(f"Selected User: {selected_user.get_full_name()}")
                    print(f"Office Manager: {office_manager.get_full_name()}")
                    print(f"Office Manager Balance: ${office_manager.balance}")
                    
                    # Check office manager balance
                    if office_manager.balance < expense_amount:
                        messages.error(
                            request, 
                            f"Insufficient balance! Office Manager's balance: ${office_manager.balance:.2f}, "
                            f"Required: ${expense_amount:.2f}"
                        )
                        print(f"\n❌ INSUFFICIENT BALANCE!")
                        return render(request, "expenses/office_expense_payment_form.html", {
                            "form": form, 
                            "expense": expense,
                            "office_manager": office_manager
                        })
                    
                    # Create payment record
                    payment = form.save(commit=False)
                    payment.expense = expense
                    payment.user = selected_user
                    payment.amount = expense_amount
                    
                    # Deduct from office manager balance
                    old_balance = office_manager.balance
                    office_manager.balance -= expense_amount
                    office_manager.balance = office_manager.balance.quantize(
                        Decimal('0.000001'), 
                        rounding=ROUND_HALF_UP
                    )
                    office_manager.save(update_fields=['balance'])
                    
                    print(f"\n✅ BALANCE DEDUCTED FROM OFFICE MANAGER")
                    print(f"Old Balance: ${old_balance}")
                    print(f"Deducted: ${expense_amount}")
                    print(f"New Balance: ${office_manager.balance}")
                    
                    # Update selected user's balance based on receive_type
                    receive_type = form.cleaned_data['receive_type']
                    
                    if receive_type == 'account':
                        # Account transfer increases user's balance
                        selected_user_obj = User.objects.select_for_update().get(id=selected_user.id)
                        old_user_balance = selected_user_obj.balance
                        selected_user_obj.balance += expense_amount
                        selected_user_obj.balance = selected_user_obj.balance.quantize(
                            Decimal('0.000001'), 
                            rounding=ROUND_HALF_UP
                        )
                        selected_user_obj.save(update_fields=['balance'])
                        
                        print(f"\n💳 ACCOUNT TRANSFER TO USER")
                        print(f"User: {selected_user_obj.get_full_name()}")
                        print(f"Old Balance: ${old_user_balance}")
                        print(f"Added: ${expense_amount}")
                        print(f"New Balance: ${selected_user_obj.balance}")
                    else:
                        # Cash payment - only deduct from office manager
                        print(f"\n💵 CASH PAYMENT - No user balance change")
                    
                    # Update expense payment status
                    expense.payment_status = "paid"
                    expense.save(update_fields=["payment_status"])
                    
                    # Save payment record
                    payment.save()
                    
                    print(f"\n✅ EXPENSE MARKED AS PAID")
                    print(f"✅ PAYMENT RECORD SAVED (ID: {payment.id})")
                    print(f"{'='*60}\n")
                    
                    messages.success(
                        request, 
                        f"✅ Payment completed successfully! ${expense_amount:.2f} deducted from office manager."
                    )
                    return redirect("accounts:office_expense_list")
                    
            except Exception as e:
                print(f"\n❌ PAYMENT ERROR: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error occurred while processing payment: {str(e)}")
        else:
            print(f"\n❌ FORM VALIDATION ERRORS:")
            print(form.errors)
            messages.error(request, "There is a problem with the form. Please fill all fields correctly.")
    
    return render(request, "expenses/office_expense_payment_form.html", {
        "form": form, 
        "expense": expense,
        "office_manager": office_manager
    })
@login_required
def office_expense_detail(request, expense_id):
    """
    View detailed information of an office expense.
    """
    if not (request.user.is_superuser or request.user.office_management):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect("accounts:dashboard")
    
    expense = get_object_or_404(
        Expense, 
        id=expense_id,
        property__isnull=True
    )
    
    # Fetch payment history
    payments = ExpensePayment.objects.filter(expense=expense).order_by('-created_at')
    
    context = {
        'expense': expense,
        'payments': payments,
    }
    
    return render(request, 'expenses/office_expense_detail.html', context)

@method_decorator(login_required, name="dispatch")
class expenselist(ListView):
    model = Expense
    template_name = "expenses/expenselist.html"
    context_object_name = "expense"
    paginate_by = 1000000

    def get_queryset(self):
        expenses = Expense.objects.all().order_by("-expense_date")
        status_filter = self.request.GET.get("status", "")
        created_by_filter = self.request.GET.get("created_by", "")
        from_date = self.request.GET.get("from_date", "")
        to_date = self.request.GET.get("to_date", "")
        if status_filter:
            expenses = expenses.filter(status=status_filter)

        if created_by_filter:
            expenses = expenses.filter(created_by_id=created_by_filter)

        if from_date:
            expenses = expenses.filter(expense_date__gte=parse_date(from_date))

        if to_date:
            expenses = expenses.filter(expense_date__lte=parse_date(to_date))

        return expenses

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Grab filters again for context
        status_filter = self.request.GET.get("status", "")
        created_by_filter = self.request.GET.get("created_by", "")
        from_date = self.request.GET.get("from_date", "")
        to_date = self.request.GET.get("to_date", "")

        context.update({
            "status_filter": status_filter,
            "created_by_filter": created_by_filter,
            "from_date": from_date,
            "to_date": to_date,
            "users": User.objects.all().order_by("short_name"),
        })
        
        filteredexpenses = self.get_queryset()
        total_expenseamount = filteredexpenses.aggregate(
            total=Sum("amount")
        )["total"] or 0

        context["total_expenseamount"] = total_expenseamount

        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance

        return context
    
@method_decorator(login_required, name="dispatch")
class managementexpenselist(ListView):
    model = Expense
    template_name = "expenses/manage_expenselist.html"
    context_object_name = "management_expenses"
    paginate_by = 100000

    def get_queryset(self):
        expenses = Expense.objects.filter(
            property__isnull=True
        ).order_by("-expense_date")

        # Get filters
        status_filter = self.request.GET.get("status", "")
        created_by_filter = self.request.GET.get("created_by", "")
        from_date = self.request.GET.get("from_date", "")
        to_date = self.request.GET.get("to_date", "")

        # Apply filters
        if status_filter:
            expenses = expenses.filter(status=status_filter)

        if created_by_filter:
            expenses = expenses.filter(created_by_id=created_by_filter)

        if from_date:
            expenses = expenses.filter(expense_date__gte=parse_date(from_date))

        if to_date:
            expenses = expenses.filter(expense_date__lte=parse_date(to_date))

        return expenses

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Grab filters again for context
        status_filter = self.request.GET.get("status", "")
        created_by_filter = self.request.GET.get("created_by", "")
        from_date = self.request.GET.get("from_date", "")
        to_date = self.request.GET.get("to_date", "")

        context.update({
            "status_filter": status_filter,
            "created_by_filter": created_by_filter,
            "from_date": from_date,
            "to_date": to_date,
            "users": User.objects.all().order_by("short_name"),
        })
        
        filtered_expenses = self.get_queryset()
        total_expense_amount = filtered_expenses.aggregate(
            total=Sum("amount")
        )["total"] or 0

        context["total_expense_amount"] = total_expense_amount

        # Total balance logic
        if self.request.user.is_superuser:
            total_investment = User.objects.aggregate(Sum("balance"))["balance__sum"] or 0
            context["total_balance"] = total_investment
        else:
            context["total_balance"] = self.request.user.balance

        return context
    




# views.py
# from decimal import Decimal
# from datetime import datetime, date
# import openpyxl

# from django.contrib import messages
# from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
# from django.db import transaction
# from django.shortcuts import redirect, render
# from django.urls import reverse_lazy
# from django.views import View

# from .forms import PropertyExcelUploadForm
# from .models import Property, User


# # ==============================
# # Excel Header (Form label) -> model field
# # ==============================
# PROPERTY_EXCEL_HEADER_MAP = {
#     "Property Name": "property_name",
#     "Description of the property": "description",
#     "Address of the property": "address",

#     "Bedrooms": "bedrooms",
#     "Bathrooms": "bathrooms",
#     "Parking": "parking",
#     "Living Area (Sq ft)": "living_area",
#     "Lot Size (Sq ft)": "lot_area",

#     "Property Type": "property_type",
#     "Year Built": "year_build",
#     "Exterior Feature": "exterior_feature",

#     "Neighborhood Demographic Profile": "neighborhood_Demographic_Profile",
#     "Neighborhood Percentage": "neighborhood_percentage",

#     "Status": "status",
#     "Auction Date": "auction_date",
#     "Auction Price ($)": "auction_price",
#     "Original listing price [Zillow/Redfine] ($)": "estimated_price",
#     "Earnest money deposit ($)": "booking_fee",
#     "URL": "url",

#     "Buying Date": "buying_date",
#     "Buying Price ($) [bp]": "buying_price",
#     "Service Cost ($) [sc]": "service_cost",
#     "Acquisition cost ($) [bp + sc]": "acquisition_cost",

#     "New listing Price ($)": "asking_price",
#     "Final Sold Price ($)": "selling_price",
#     "Selling Date": "selling_date",

#     "Listed By Email (optional)": "listed_by_email",
# }

# ALLOWED_FIELD_HEADERS = set(PROPERTY_EXCEL_HEADER_MAP.values()) | {"listed_by_email"}


# # ✅ Must match your STATUS_CHOICES keys
# VALID_STATUSES = {
#     "wishlist",
#     "failed_to_bought",
#     "move_to_next_option",
#     "bought",
#     "ready_to_sell",
#     "sold",
# }


# def _clean_str(v):
#     if v is None:
#         return ""
#     s = str(v).strip()
#     if s.lower() in ("none", "null", "nan"):
#         return ""
#     return s


# def _to_decimal(val, default=None):
#     s = _clean_str(val)
#     if s == "":
#         return default
#     try:
#         return Decimal(s)
#     except Exception:
#         return default


# def _to_int(val, default=None):
#     s = _clean_str(val)
#     if s == "":
#         return default
#     try:
#         return int(float(s))
#     except Exception:
#         return default


# def _to_float(val, default=None):
#     s = _clean_str(val)
#     if s == "":
#         return default
#     try:
#         return float(s)
#     except Exception:
#         return default


# def _to_date(val, default=None):
#     if val is None or val == "":
#         return default

#     # Excel datetime
#     if hasattr(val, "date"):
#         try:
#             return val.date()
#         except Exception:
#             pass

#     s = _clean_str(val)
#     if s == "":
#         return default

#     for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
#         try:
#             return datetime.strptime(s, fmt).date()
#         except Exception:
#             continue

#     return default


# class IsSuperUserOrPropertyMixin(UserPassesTestMixin):
#     def test_func(self):
#         u = self.request.user
#         return u.is_authenticated and (u.is_superuser or getattr(u, "is_property", False) or u.is_staff)


# class PropertyExcelUploadView(LoginRequiredMixin, IsSuperUserOrPropertyMixin, View):
#     template_name = "property_excel_upload.html"
#     success_url = reverse_lazy("accounts:property_list")

#     def get(self, request):
#         return render(request, self.template_name, {"form": PropertyExcelUploadForm()})

#     def post(self, request):
#         form = PropertyExcelUploadForm(request.POST, request.FILES)
#         if not form.is_valid():
#             return render(request, self.template_name, {"form": form})

#         f = form.cleaned_data["file"]
#         if not f.name.lower().endswith(".xlsx"):
#             messages.error(request, "Please upload a valid .xlsx file.")
#             return render(request, self.template_name, {"form": form})

#         try:
#             wb = openpyxl.load_workbook(f, data_only=True)
#         except Exception as e:
#             messages.error(request, f"Excel read failed: {e}")
#             return render(request, self.template_name, {"form": form})

#         # ✅ only property required
#         if "property" not in wb.sheetnames:
#             messages.error(request, "Required sheet missing: property")
#             return render(request, self.template_name, {"form": form})

#         ws_prop = wb["property"]
#         ws_inv = wb["investments"] if "investments" in wb.sheetnames else None  # optional

#         # -------------------------
#         # Validate property headers
#         # -------------------------
#         raw_headers = [(_clean_str(c.value)) for c in next(ws_prop.iter_rows(min_row=1, max_row=1))]

#         mapped_headers = []
#         unknown = []
#         for h in raw_headers:
#             if not h:
#                 mapped_headers.append("")
#                 continue

#             if h in PROPERTY_EXCEL_HEADER_MAP:
#                 mapped_headers.append(PROPERTY_EXCEL_HEADER_MAP[h])
#             else:
#                 # allow direct field names
#                 mapped_headers.append(h)
#                 if h not in ALLOWED_FIELD_HEADERS:
#                     unknown.append(h)

#         if unknown:
#             messages.error(request, f"Unknown column header(s): {', '.join(unknown)}")
#             return render(request, self.template_name, {"form": form})

#         # -------------------------
#         # Read investments (optional)
#         # -------------------------
#         inv_by_property = {}
#         inv_dates_by_property = {}

#         if ws_inv:
#             inv_headers = [(_clean_str(c.value)) for c in next(ws_inv.iter_rows(min_row=1, max_row=1))]

#             for row in ws_inv.iter_rows(min_row=2, values_only=True):
#                 if not any(v is not None and v != "" for v in row):
#                     continue

#                 d = dict(zip(inv_headers, row))
#                 p_name = _clean_str(d.get("property_name"))
#                 if not p_name:
#                     continue

#                 email = _clean_str(d.get("user_email"))
#                 if not email:
#                     continue

#                 user = User.objects.filter(email__iexact=email).first()
#                 if not user:
#                     messages.error(request, f"User not found for email: {email}")
#                     return render(request, self.template_name, {"form": form})

#                 invest_amount = _to_decimal(d.get("invest_amount"), default=Decimal("0"))
#                 if invest_amount <= 0:
#                     continue

#                 is_fixed_raw = d.get("is_fixed")
#                 if isinstance(is_fixed_raw, bool):
#                     is_fixed = is_fixed_raw
#                 else:
#                     is_fixed = _clean_str(is_fixed_raw).lower() in ("true", "yes", "1")

#                 sequence = _to_int(d.get("sequence"), default=1) or 1
#                 inv_date = _to_date(d.get("investment_date"), default=None)

#                 inv_by_property.setdefault(p_name, []).append({
#                     "user_id": user.id,
#                     "invest_amount": invest_amount,
#                     "is_fixed": is_fixed,
#                     "sequence": sequence,
#                 })

#                 if inv_date:
#                     inv_dates_by_property.setdefault(p_name, []).append({
#                         "user_id": user.id,
#                         "sequence": sequence,
#                         "date": inv_date,
#                     })

#         created_count = 0
#         updated_count = 0

#         try:
#             with transaction.atomic():
#                 for row in ws_prop.iter_rows(min_row=2, values_only=True):
#                     if not any(v is not None and v != "" for v in row):
#                         continue

#                     prop_data = dict(zip(mapped_headers, row))
#                     property_name = _clean_str(prop_data.get("property_name"))
#                     if not property_name:
#                         continue

#                     # ✅ status blank => wishlist
#                     raw_status = _clean_str(prop_data.get("status")).lower()
#                     status = raw_status or "wishlist"

#                     # ✅ validate against your STATUS_CHOICES keys
#                     if status not in VALID_STATUSES:
#                         raise ValueError(
#                             f"Invalid status '{status}' for property '{property_name}'. "
#                             f"Allowed: {', '.join(sorted(VALID_STATUSES))}"
#                         )

#                     # optional listed_by from email
#                     listed_by = None
#                     listed_by_email = _clean_str(prop_data.get("listed_by_email"))
#                     if listed_by_email:
#                         listed_by = User.objects.filter(email__iexact=listed_by_email).first()

#                     defaults = {
#                         "description": prop_data.get("description"),
#                         "status": status,
#                         "address": prop_data.get("address"),
#                         "url": (_clean_str(prop_data.get("url")) or None),

#                         "auction_date": _to_date(prop_data.get("auction_date")),
#                         "auction_price": _to_decimal(prop_data.get("auction_price")),
#                         "estimated_price": _to_decimal(prop_data.get("estimated_price")),
#                         "booking_fee": _to_decimal(prop_data.get("booking_fee")),

#                         "buying_date": _to_date(prop_data.get("buying_date")),
#                         "buying_price": _to_decimal(prop_data.get("buying_price")),
#                         "service_cost": _to_decimal(prop_data.get("service_cost")),
#                         "acquisition_cost": _to_decimal(prop_data.get("acquisition_cost")),

#                         "asking_price": _to_decimal(prop_data.get("asking_price")),
#                         "selling_price": _to_decimal(prop_data.get("selling_price")),
#                         "selling_date": _to_date(prop_data.get("selling_date")),

#                         "bedrooms": _to_int(prop_data.get("bedrooms")),
#                         "bathrooms": _to_float(prop_data.get("bathrooms")),
#                         "living_area": _to_int(prop_data.get("living_area")),
#                         "lot_area": _to_int(prop_data.get("lot_area")),
#                         "parking": _to_int(prop_data.get("parking")),
#                         "year_build": _to_int(prop_data.get("year_build")),

#                         "property_type": _clean_str(prop_data.get("property_type")) or "single_family",
#                         "exterior_feature": _clean_str(prop_data.get("exterior_feature")) or "brick",
#                         "neighborhood_Demographic_Profile": _clean_str(prop_data.get("neighborhood_Demographic_Profile")) or "White (Non-Hispanic)",
#                         "neighborhood_percentage": _to_int(prop_data.get("neighborhood_percentage")),
#                     }

#                     # ✅ update existing by property_name, else create new
#                     prop, created = Property.objects.update_or_create(
#                         property_name=property_name,
#                         defaults=defaults
#                     )

#                     if listed_by:
#                         prop.listed_by = listed_by
#                         prop.save(update_fields=["listed_by"])

#                     if created:
#                         created_count += 1
#                     else:
#                         updated_count += 1

#                     # ✅ optional deduction
#                     investments_list = inv_by_property.get(property_name, [])
#                     investment_dates_list = inv_dates_by_property.get(property_name, [])

#                     total_cost = (prop.buying_price or Decimal("0")) + (prop.service_cost or Decimal("0"))
#                     should_run_deduction = (total_cost > 0) and bool(investments_list)

#                     if should_run_deduction:
#                         user_ids = list({inv["user_id"] for inv in investments_list})
#                         contributors = User.objects.filter(id__in=user_ids, is_active=True)
#                         prop.contributors.add(*contributors)

#                         ok = prop.deduct_property_costs_with_multiple_investments(
#                             investments_list=investments_list,
#                             investment_dates_list=investment_dates_list or None,
#                         )
#                         if not ok:
#                             raise ValueError(f"Contribution deduction failed for: {property_name}")

#         except Exception as e:
#             messages.error(request, f"Upload failed: {e}")
#             return render(request, self.template_name, {"form": form})

#         messages.success(request, f"Excel processed. Created: {created_count}, Updated: {updated_count}")
#         return redirect(self.success_url)


import re
from decimal import Decimal
from datetime import datetime

import openpyxl

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views import View

from .forms import PropertyExcelUploadForm
from .models import Property, User


# ===============================
# Exact Excel Header Mapping
# ===============================
HEADER_MAP = {
    "Property Name": "property_name",
    "Address": "address",
    "Address ": "address",

    "Neighborhood": "neighborhood_Demographic_Profile",
    "Percentage": "neighborhood_percentage",

    "Listing price": "estimated_price",
    "Auction Price": "auction_price",
    "Auction Price ": "auction_price",

    "LP-AP": "__ignore__",

    "Bedrooms": "bedrooms",
    "Bathrooms": "bathrooms",
    "Living Area": "living_area",
    "Lot Size": "lot_area",

    "URL": "url",
    "Parking": "parking",
    "Property Type": "property_type",
    "Year Built": "year_build",
    "Exterior": "exterior_feature",
    "Description": "description",

    "Status": "status",
    "Auction Date": "auction_date",
    "Deposit": "booking_fee",
    "Deposit ": "booking_fee",

    "Buying Date": "buying_date",
    "Buying Price": "buying_price",
    "Service Cost": "service_cost",
    "Acquisition Cost": "acquisition_cost",
    "New listing Price": "asking_price",
    "Final Sold Price": "selling_price",
    "Selling Date": "selling_date",
    "Listed By (Email)": "listed_by_email",
}


# ===============================
# Status & Neighbor Mapping
# ===============================
STATUS_MAP = {
    "1": "wishlist",
    "2": "failed_to_bought",
    "3": "move_to_next_option",
    "4": "bought",
    "5": "ready_to_sell",
    "6": "sold",
}

NEIGHBOR_MAP = {
    "1": "White (Non-Hispanic)",
    "2": "Black or African American",
    "3": "Asian",
    "4": "Hispanic or Latino",
}


# ===============================
# Helpers
# ===============================
def clean(val):
    if val is None:
        return ""
    return str(val).strip()


def has_value(val):
    if val is None:
        return False
    if isinstance(val, str) and val.strip() == "":
        return False
    return True


def parse_decimal(val):
    if not has_value(val):
        return None

    if isinstance(val, (int, float)):
        return Decimal(str(val))

    s = str(val).replace("$", "").replace(",", "").strip().lower()

    if s.endswith("k"):
        s = str(Decimal(s[:-1]) * 1000)

    try:
        return Decimal(s)
    except:
        return None


def parse_int(val):
    if not has_value(val):
        return None
    try:
        return int(float(val))
    except:
        return None


def parse_float(val):
    if not has_value(val):
        return None
    try:
        return float(val)
    except:
        return None


def parse_date(val):
    if hasattr(val, "date"):
        return val.date()

    if not has_value(val):
        return None

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val), fmt).date()
        except:
            continue
    return None


def normalize_status(val):
    if not has_value(val):
        return STATUS_MAP["1"]

    if isinstance(val, (int, float)):
        val = str(int(val))

    val = str(val).lower().strip()

    if val in STATUS_MAP:
        return STATUS_MAP[val]

    return val


def normalize_neighbor(val):
    if not has_value(val):
        return NEIGHBOR_MAP["1"]

    if isinstance(val, (int, float)):
        val = str(int(val))

    val = str(val).strip()

    if val in NEIGHBOR_MAP:
        return NEIGHBOR_MAP[val]

    return val


# ===============================
# Permission
# ===============================
class IsSuperUserOrPropertyMixin(UserPassesTestMixin):
    def test_func(self):
        u = self.request.user
        return u.is_authenticated and (u.is_superuser or getattr(u, "is_property", False))


# ===============================
# FINAL VIEW
# ===============================
class PropertyExcelUploadView(LoginRequiredMixin, IsSuperUserOrPropertyMixin, View):
    template_name = "property_excel_upload.html"
    success_url = reverse_lazy("accounts:property_list")

    def get(self, request):
        return render(request, self.template_name, {"form": PropertyExcelUploadForm()})

    def post(self, request):
        form = PropertyExcelUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {"form": form})

        file = form.cleaned_data["file"]

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        raw_headers = [clean(c.value) for c in next(ws.iter_rows(max_row=1))]
        headers = [HEADER_MAP.get(h, h) for h in raw_headers]

        created = 0
        updated = 0

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(has_value(v) for v in row):
                    continue

                data = dict(zip(headers, row))
                data.pop("__ignore__", None)

                property_name = clean(data.get("property_name"))
                if not property_name:
                    continue

                prop, was_created = Property.objects.get_or_create(
                    property_name=property_name
                )

                # Always set defaults
                prop.status = normalize_status(data.get("status"))
                prop.neighborhood_Demographic_Profile = normalize_neighbor(
                    data.get("neighborhood_Demographic_Profile")
                )

                # Update only if value exists
                for field in [
                    "description", "address", "url",
                ]:
                    if has_value(data.get(field)):
                        setattr(prop, field, clean(data.get(field)))

                if has_value(data.get("auction_price")):
                    prop.auction_price = parse_decimal(data.get("auction_price"))

                if has_value(data.get("estimated_price")):
                    prop.estimated_price = parse_decimal(data.get("estimated_price"))

                if has_value(data.get("booking_fee")):
                    prop.booking_fee = parse_decimal(data.get("booking_fee"))

                if has_value(data.get("buying_price")):
                    prop.buying_price = parse_decimal(data.get("buying_price"))

                if has_value(data.get("service_cost")):
                    prop.service_cost = parse_decimal(data.get("service_cost"))

                if has_value(data.get("acquisition_cost")):
                    prop.acquisition_cost = parse_decimal(data.get("acquisition_cost"))

                if has_value(data.get("asking_price")):
                    prop.asking_price = parse_decimal(data.get("asking_price"))

                if has_value(data.get("selling_price")):
                    prop.selling_price = parse_decimal(data.get("selling_price"))

                if has_value(data.get("bedrooms")):
                    prop.bedrooms = parse_int(data.get("bedrooms"))

                if has_value(data.get("bathrooms")):
                    prop.bathrooms = parse_float(data.get("bathrooms"))

                if has_value(data.get("living_area")):
                    prop.living_area = parse_int(data.get("living_area"))

                if has_value(data.get("lot_area")):
                    prop.lot_area = parse_int(data.get("lot_area"))

                if has_value(data.get("parking")):
                    prop.parking = parse_int(data.get("parking"))

                if has_value(data.get("year_build")):
                    prop.year_build = parse_int(data.get("year_build"))

                if has_value(data.get("neighborhood_percentage")):
                    prop.neighborhood_percentage = parse_int(data.get("neighborhood_percentage"))

                if has_value(data.get("auction_date")):
                    prop.auction_date = parse_date(data.get("auction_date"))

                if has_value(data.get("buying_date")):
                    prop.buying_date = parse_date(data.get("buying_date"))

                if has_value(data.get("selling_date")):
                    prop.selling_date = parse_date(data.get("selling_date"))

                listed_email = clean(data.get("listed_by_email"))
                if listed_email:
                    user = User.objects.filter(email__iexact=listed_email).first()
                    if user:
                        prop.listed_by = user

                prop.save()

                if was_created:
                    created += 1
                else:
                    updated += 1

        messages.success(request, f"Upload done. Created: {created}, Updated: {updated}")
        return redirect(self.success_url)
    